from datetime import date, timedelta
from calendar import monthrange
from typing import Literal
from dotenv import load_dotenv
from pydantic import BaseModel, Field
from langchain_openai import ChatOpenAI
from openai import OpenAI
import os
import re
from xlsm_safe import atomic_save_workbook

try:
    from markitdown import MarkItDown
except ImportError:
    MarkItDown = None

load_dotenv()

DOSSIER_FACTURES = "factures"

class DocumentInfo(BaseModel):
    type_document: Literal["facture", "bon_livraison"] | None = None
    numero_facture: str | None = None
    numero_bon_livraison: str | None = None
    date_emission: date | None = None
    date_livraison: date | None = None
    date_paiement_prevue: date | None = None
    montant_total: float | None = None
    nom_fournisseur: str | None = None
    bons_livraisons: list[str] = Field(default_factory=list)
    conditions_paiement: str | None = None
    prix_HT_5_5pct: float | None = None
    prix_HT_10pct: float | None = None
    prix_HT_20pct: float | None = None

def build_apim_headers(feature: str, api_key: str | None = None) -> dict[str, str]:
    return {"api-key": api_key or "", "owner": "HAMILTON", "feature": feature}

APIM_OPENAI_BASE_URL = os.getenv("APIM_OPENAI_BASE_URL")
APIM_OPENAI_API_KEY = os.getenv("APIM_OPENAI_API_KEY")
api_key = APIM_OPENAI_API_KEY
feature = "my-feature"

llm = None
llm_client = None
if APIM_OPENAI_BASE_URL and APIM_OPENAI_API_KEY:
    llm = ChatOpenAI(
        model="gpt-5.1-2025-11-13",
        api_key=api_key,  # type: ignore[arg-type]
        base_url=APIM_OPENAI_BASE_URL,
        default_headers=build_apim_headers(feature=feature, api_key=api_key),
        use_responses_api=False,
        streaming=False,
        reasoning_effort="low",
        temperature=0,
        max_retries=3,
        max_completion_tokens=1024,
        verbose=True,
    ).with_structured_output(
        DocumentInfo,
        method="json_schema",
        strict=False,
        include_raw=False,
    )

    llm_client = OpenAI(
        api_key=api_key,
        base_url=APIM_OPENAI_BASE_URL,
        default_headers=build_apim_headers(feature=feature, api_key=api_key),
    )
else:
    print("[WARN] APIM_OPENAI_* non configure: endpoints IA indisponibles, API edition reste active.")

if MarkItDown is not None and llm_client is not None:
    md = MarkItDown(enable_plugins=True, llm_client=llm_client, llm_model="gpt-5.1-2025-11-13")
else:
    md = None

print("Initialisation terminée. Prêt à traiter les documents.")

def load_pdf_text(filepath: str) -> str:
    if md is None:
        raise RuntimeError(
            "Le package 'markitdown' n'est pas installe. "
            "Installez-le pour activer l'extraction PDF (pip install markitdown)."
        )
    return md.convert(filepath).text_content

def extract_date_candidates(text: str) -> list[date]:
    out, seen = [], set()
    for d, m, y in re.findall(r"\b(\d{2})[/-](\d{2})[/-](\d{4})\b", text):
        try:
            dt = date(int(y), int(m), int(d))
            if 2020 <= dt.year <= 2100 and dt not in seen:
                seen.add(dt)
                out.append(dt)
        except ValueError:
            pass
    for y, m, d in re.findall(r"\b(\d{4})-(\d{2})-(\d{2})\b", text):
        try:
            dt = date(int(y), int(m), int(d))
            if 2020 <= dt.year <= 2100 and dt not in seen:
                seen.add(dt)
                out.append(dt)
        except ValueError:
            pass
    return sorted(out)

def extract_date_from_filename(filename: str) -> date | None:
    m = re.search(r"(20\d{2})(\d{2})(\d{2})", filename)
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None

def parse_date_string(value: str | None) -> date | None:
    if not value:
        return None
    value = value.strip()
    m = re.fullmatch(r"(\d{2})[/-](\d{2})[/-](\d{4})", value)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", value)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None

def is_supported_date(value: date | None) -> bool:
    return isinstance(value, date) and 2020 <= value.year <= 2100

def extract_labeled_date(text: str, patterns: list[str]) -> date | None:
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if not match:
            continue
        parsed = parse_date_string(match.group(1))
        if is_supported_date(parsed):
            return parsed
    return None

def choose_best_date(
    raw_value: date | str | None,
    *,
    text: str,
    filename: str,
    label_patterns: list[str],
    fallback: date | None = None,
    prefer_last_candidate: bool = False,
) -> date | None:
    current = raw_value if isinstance(raw_value, date) else parse_date_string(raw_value)
    candidates = extract_date_candidates(text)
    explicit = extract_labeled_date(text, label_patterns)
    file_date = extract_date_from_filename(filename)

    if explicit:
        return explicit
    if is_supported_date(current):
        return current
    if isinstance(current, date):
        for candidate in candidates:
            if candidate.day == current.day and candidate.month == current.month:
                return candidate
    if fallback and is_supported_date(fallback):
        return fallback
    if file_date and is_supported_date(file_date):
        return file_date
    if not candidates:
        return None
    return candidates[-1] if prefer_last_candidate else candidates[0]

def infer_due_date(emission: date | None, text: str) -> date | None:
    if not emission:
        return None
    lower = text.lower()
    m = re.search(r"(\d{1,3})\s*jours?", lower)
    delay = int(m.group(1)) if m else None
    if delay is None:
        return None
    if "fin de mois" in lower:
        eom = date(emission.year, emission.month, monthrange(emission.year, emission.month)[1])
        return eom + timedelta(days=delay)
    return emission + timedelta(days=delay)

def reconcile_due_date_with_terms(
    due: date | None,
    *,
    emission: date | None,
    text: str,
) -> date | None:
    inferred_due = infer_due_date(emission, text)
    if not inferred_due:
        return due
    if not due:
        return inferred_due
    return due if due == inferred_due else inferred_due

def clean_bl_number(value: str) -> str:
    v = str(value).upper().strip()
    v = re.sub(r"^BL\s*N[°O]?\s*", "", v)
    v = re.sub(r"^BON\s+DE\s+LIVRAISON\s*N[°O]?\s*", "", v)
    v = re.sub(r"^AR\s*CDE\s*N[°O]?\s*", "", v)
    v = re.sub(r"\s+", "", v)
    return v

def clean_invoice_number(value: str) -> str:
    v = str(value).strip().upper()
    v = re.sub(r"^(FACTURE|FAC)\s*N[°O]?\s*", "", v)
    v = re.sub(r"\s+", "", v)
    return v.strip("-:/")

def normalize_bl_list(values: list[str]) -> list[str]:
    cleaned, seen = [], set()
    for v in values or []:
        c = clean_bl_number(v)
        if c and c not in seen:
            seen.add(c)
            cleaned.append(c)
    return cleaned

def normalize_supplier_name(
    value: str | None,
    fournisseur_patterns: dict[str, list[str]] | None = None,
) -> str | None:
    """
    Normalise un nom de fournisseur brut vers son identifiant interne.
    Si fournisseur_patterns est fourni (dict {id: [patterns]}), l'utilise.
    Sinon, repli sur les 3 fournisseurs historiques codés en dur.
    """
    if not value:
        return None
    raw = re.sub(r"\s+", " ", value.strip().lower())

    if fournisseur_patterns:
        for fournisseur_id, patterns in fournisseur_patterns.items():
            for pattern in patterns:
                if pattern.lower() in raw:
                    return fournisseur_id
        # Tentative directe sur l'identifiant lui-même
        raw_upper = raw.upper().replace(" ", "")
        if raw_upper in fournisseur_patterns:
            return raw_upper
        return None

    # Repli historique
    raw_upper = re.sub(r"\s+", " ", value.upper().strip())
    if "AMBELYS" in raw_upper:
        return "AMBELYS"
    if "SYSCO" in raw_upper:
        return "SYSCO"
    if "TERREAZUR" in raw_upper or "TERRE AZUR" in raw_upper:
        return "TERREAZUR"
    return None

def infer_supplier_name(
    text: str,
    filename: str,
    fournisseur_patterns: dict[str, list[str]] | None = None,
) -> str | None:
    haystack = f"{filename}\n{text}".lower()
    if fournisseur_patterns:
        for fournisseur_id, patterns in fournisseur_patterns.items():
            for pattern in patterns:
                if pattern.lower() in haystack:
                    return fournisseur_id
        return None
    if "ambelys" in haystack:
        return "AMBELYS"
    if "sysco" in haystack:
        return "SYSCO"
    if "terre azur" in haystack or "terreazur" in haystack:
        return "TERREAZUR"
    return None

def extract_invoice_number(text: str, filename: str) -> str | None:
    patterns = [
        r"\bnum[eé]ro\s+de\s+facture\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-\/]*)",
        r"\bfacture\s*n[°o]?\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-\/]*)",
        r"\bfacture\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-\/]*)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return clean_invoice_number(match.group(1))

    fname = os.path.splitext(os.path.basename(filename))[0].upper()
    match = re.search(r"(?:^|[-_\s])((?:FAC[-_\/]?|F)\d{4,})\b", fname)
    if match:
        return clean_invoice_number(match.group(1))
    return None

def extract_payment_terms(text: str) -> str | None:
    patterns = [
        r"conditions?\s+de\s+r[eè]glement\s*[:\-]?\s*([^\n\r]+)",
        r"modalit[eé]s?\s+de\s+paiement\s*[:\-]?\s*([^\n\r]+)",
        r"\b(payable\s+[^\n\r]{0,80})",
        r"\b(\d{1,3}\s*jours?(?:\s+fin\s+de\s+mois)?)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            value = re.sub(r"\s+", " ", match.group(1)).strip(" .:-")
            if value:
                return value
    return None

def extract_referenced_bl_numbers(text: str) -> list[str]:
    pattern = (
        r"\b(?:bon\s+de\s+livraison|bl|ar\s*cde)"
        r"\s*n[°o]?\s*[:\-]?\s*([A-Z0-9\-\/]+)"
    )
    found = [match.group(1).upper() for match in re.finditer(pattern, text, re.IGNORECASE)]
    return normalize_bl_list(found)

def classify_document(text: str, filename: str) -> str:
    """
    Classifie un document en 'facture' ou 'bon_livraison'.

    Stratégie en 3 passes :
      1. Règles déterministes sur le nom de fichier (patterns fournisseurs connus)
      2. Règles déterministes sur le contenu textuel (marqueurs forts)
      3. Scoring pondéré sur le contenu (arbitrage en cas d'ambiguïté)
    """
    lower = text.lower()
    # On analyse les 12 000 premiers caractères pour couvrir les en-têtes longs
    head = lower[:12_000]
    f = filename.lower()
    fname = os.path.splitext(os.path.basename(f))[0]  # nom sans extension

    # ------------------------------------------------------------------
    # PASSE 1 — Patterns déterministes sur le nom de fichier
    # ------------------------------------------------------------------

    # AMBELYS : préfixe numérique + lettre de type
    #   C = commande / bon de livraison  (ex: 01_AMBELYS-C215075)
    #   F = facture                      (ex: 01_F826802)
    if re.search(r"\bambelys", f):
        if re.search(r"[-_]c\d{4,}", fname):          # -C215075
            return "bon_livraison"
        if re.search(r"[-_]f\d{4,}", fname):          # -F826802 ou _F826802
            return "facture"

    # SYSCO : patterns de nommage connus
    #   BL / livraison / cde dans le nom → BL
    #   F / FAC / facture dans le nom → facture
    if re.search(r"\bsysco", f):
        if re.search(r"[-_ ](bl|bon|livraison|cde|cmd)\b", fname):
            return "bon_livraison"
        if re.search(r"[-_ ](f|fac|facture)\b", fname):
            return "facture"

    # TERREAZUR / TERRE AZUR
    if re.search(r"terre\s*azur", f):
        if re.search(r"[-_ ](bl|bon|livraison|cde)\b", fname):
            return "bon_livraison"
        if re.search(r"[-_ ](f|fac|facture)\b", fname):
            return "facture"

    # Patterns génériques dans le nom de fichier
    if re.search(r"\b(bon[_\- ]de[_\- ]livraison|bon[_\- ]livraison)\b", fname):
        return "bon_livraison"
    if re.search(r"\bfacture\b", fname):
        return "facture"

    # ------------------------------------------------------------------
    # PASSE 2 — Marqueurs forts dans le contenu (quasi-certains)
    # ------------------------------------------------------------------

    # Marqueurs BL très forts (présents uniquement sur des BL)
    BL_STRONG = [
        r"\ba\s+livrer\s+le\b",          # "à livrer le"
        r"\bbon\s+de\s+livraison\b",
        r"\bbon\s+livraison\b",
        r"\bdate\s+de\s+livraison\b",
        r"\bar\s*cde\s*n[°o]?\s*\d+",   # AR CDE N°...
        r"\bcommande\s+n[°o]?\s*\d+",
        r"\bréférence\s+commande\b",
    ]
    bl_strong_hits = sum(1 for p in BL_STRONG if re.search(p, head))

    # Marqueurs facture forts
    INVOICE_STRONG = [
        r"\bdate\s+de\s+facture\b",
        r"\bnum[eé]ro\s+de\s+facture\b",
        r"\bfacture\s+n[°o]?\s*\d+",
        r"\b[eé]ch[eé]ance\b",
        r"\bconditions\s+de\s+r[eè]glement\b",
        r"\bnet\s+[àa]\s+payer\b",
        r"\btotal\s+ttc\b",
    ]
    inv_strong_hits = sum(1 for p in INVOICE_STRONG if re.search(p, head))

    # Si un type domine clairement → décision immédiate
    if bl_strong_hits >= 2 and inv_strong_hits == 0:
        return "bon_livraison"
    if inv_strong_hits >= 2 and bl_strong_hits == 0:
        return "facture"

    # ------------------------------------------------------------------
    # PASSE 3 — Scoring pondéré (arbitrage)
    # ------------------------------------------------------------------
    invoice_score = 0
    delivery_score = 0

    # Signaux BL
    if re.search(r"\bar\s*cde\s*n[°o]?\s*\d+", head):
        delivery_score += 10
    if re.search(r"\ba\s+livrer\s+le\b", head):
        delivery_score += 8
    if re.search(r"\bbon\s+de\s+livraison\b", head):
        delivery_score += 8
    if re.search(r"\bbon\s+livraison\b", head):
        delivery_score += 6
    if re.search(r"\bdate\s+de\s+livraison\b", head):
        delivery_score += 5
    if re.search(r"\bbl\s+n[°o]?\s*\d+", head):
        delivery_score += 6
    if re.search(r"\bcommande\s+n[°o]?\s*\d+", head):
        delivery_score += 4
    if re.search(r"\bquantit[eé]\s+(livr[eé]e?|command[eé]e?)\b", head):
        delivery_score += 4
    # Nom de fichier BL
    if re.search(r"\b(bl|livraison|cde|cmd|bon)\b", fname):
        delivery_score += 3

    # Signaux facture
    if re.search(r"\bfacture\s+n[°o]?\s*\d+", head):
        invoice_score += 10
    if re.search(r"\bdate\s+de\s+facture\b", head):
        invoice_score += 8
    if re.search(r"\b[eé]ch[eé]ance\b", head):
        invoice_score += 6
    if re.search(r"\bnet\s+[àa]\s+payer\b", head):
        invoice_score += 6
    if re.search(r"\btotal\s+ttc\b", head):
        invoice_score += 5
    if re.search(r"\btva\b", head):
        invoice_score += 3
    if re.search(r"\bttc\b", head):
        invoice_score += 2
    if re.search(r"\bconditions\s+de\s+r[eè]glement\b", head):
        invoice_score += 4
    # Nom de fichier facture
    if re.search(r"\b(fac|facture|invoice)\b", fname):
        invoice_score += 3

    # Intégration des hits forts dans le score
    delivery_score += bl_strong_hits * 3
    invoice_score  += inv_strong_hits * 3

    # En cas d'égalité parfaite → on regarde si "facture" apparaît dans le texte
    # mais uniquement en dehors d'une phrase de référence BL
    if delivery_score == invoice_score:
        # "facture" mentionné comme référence dans un BL ne compte pas
        facture_ref = re.search(
            r"(bon de livraison|bl n[°o]?).{0,200}facture|facture.{0,200}(bon de livraison|bl n[°o]?)",
            head
        )
        if not facture_ref and "facture" in head:
            invoice_score += 1

    return "bon_livraison" if delivery_score > invoice_score else "facture"

def normalize_invoice_dates(data: dict, text: str, filename: str) -> dict:
    candidates = extract_date_candidates(text)

    emission = choose_best_date(
        data.get("date_emission"),
        text=text,
        filename=filename,
        label_patterns=[
            r"date\s+de\s+facture\s*[:\-]?\s*(\d{2}[/-]\d{2}[/-]\d{4}|\d{4}-\d{2}-\d{2})",
            r"date\s+d[’']?[eé]mission\s*[:\-]?\s*(\d{2}[/-]\d{2}[/-]\d{4}|\d{4}-\d{2}-\d{2})",
            r"facture\s+du\s+(\d{2}[/-]\d{2}[/-]\d{4}|\d{4}-\d{2}-\d{2})",
        ],
    )
    data["date_emission"] = emission

    due = choose_best_date(
        data.get("date_paiement_prevue"),
        text=text,
        filename=filename,
        label_patterns=[
            r"[eé]ch[eé]ance\s*[:\-]?\s*(\d{2}[/-]\d{2}[/-]\d{4}|\d{4}-\d{2}-\d{2})",
            r"date\s+de\s+paiement\s*[:\-]?\s*(\d{2}[/-]\d{2}[/-]\d{4}|\d{4}-\d{2}-\d{2})",
        ],
        fallback=infer_due_date(emission, text),
        prefer_last_candidate=True,
    )
    if due is None:
        later = [d for d in candidates if emission and d >= emission]
        due = later[-1] if later else None
    due = reconcile_due_date_with_terms(due, emission=emission, text=text)
    data["date_paiement_prevue"] = due
    return data

def finalize_document_data(
    data: dict,
    text: str,
    filename: str,
    predicted_type: str,
    fournisseur_patterns: dict[str, list[str]] | None = None,
) -> dict:
    data = dict(data)
    data["type_document"] = predicted_type
    data["fichier_source"] = filename

    data["nom_fournisseur"] = normalize_supplier_name(
        data.get("nom_fournisseur"),
        fournisseur_patterns=fournisseur_patterns,
    )
    if data["nom_fournisseur"] is None:
        data["nom_fournisseur"] = infer_supplier_name(
            text,
            filename,
            fournisseur_patterns=fournisseur_patterns,
        )

    if predicted_type == "facture":
        if not data.get("numero_facture"):
            data["numero_facture"] = extract_invoice_number(text, filename)

        data = normalize_invoice_dates(data, text=text, filename=filename)

        if not data.get("conditions_paiement"):
            data["conditions_paiement"] = extract_payment_terms(text)

        merged_bls = normalize_bl_list(
            [*(data.get("bons_livraisons") or []), *extract_referenced_bl_numbers(text)]
        )
        data["bons_livraisons"] = merged_bls

    if predicted_type == "bon_livraison":
        if not data.get("numero_bon_livraison"):
            patterns = [
                r"\bar\s*cde\s*[n°o:\- ]+\s*([A-Z0-9\-\/]+)",
                r"\bbon de livraison\s*[n°o:\- ]+\s*([A-Z0-9\-\/]+)",
                r"\bbl\s*[n°o:\- ]+\s*([A-Z0-9\-\/]+)",
            ]
            for p in patterns:
                m = re.search(p, text, re.IGNORECASE)
                if m:
                    data["numero_bon_livraison"] = clean_bl_number(m.group(1))
                    break

        data["date_livraison"] = choose_best_date(
            data.get("date_livraison"),
            text=text,
            filename=filename,
            label_patterns=[
                r"date\s+de\s+livraison\s*[:\-]?\s*(\d{2}[/-]\d{2}[/-]\d{4}|\d{4}-\d{2}-\d{2})",
                r"[àa]\s+livrer\s+le\s+(\d{2}[/-]\d{2}[/-]\d{4}|\d{4}-\d{2}-\d{2})",
                r"livraison\s+du\s+(\d{2}[/-]\d{2}[/-]\d{4}|\d{4}-\d{2}-\d{2})",
            ],
        )

        data["bons_livraisons"] = normalize_bl_list(data.get("bons_livraisons", []))

    if isinstance(data.get("bons_livraisons"), list):
        data["bons_livraisons"] = normalize_bl_list(data["bons_livraisons"])

    if data.get("numero_bon_livraison"):
        data["numero_bon_livraison"] = clean_bl_number(data["numero_bon_livraison"])

    return data

def build_prompt(document_type: str, text: str, fournisseur_ids: list[str] | None = None) -> str:
    ids = fournisseur_ids or ["SYSCO", "AMBELYS", "TERREAZUR"]
    fournisseurs_str = ", ".join(ids)
    if document_type == "bon_livraison":
        return f"""
Extrais uniquement les informations du bon de livraison.
Retourne un JSON conforme au schéma.

Règles:
- type_document = bon_livraison
- Si une valeur est absente, retourne null
- numero_bon_livraison = numéro principal du BL
- date_livraison = date de livraison du BL (entre 2020 et 2100 sinon null)
- n'invente jamais une année différente de celle écrite dans le document
- prix_HT_5_5pct, prix_HT_10pct, prix_HT_20pct = montants HT de CE bon de livraison selon le taux de TVA applicable
- normalise les montants en nombres décimaux
- nom_fournisseur doit être exactement une des valeurs suivantes : {fournisseurs_str}

Texte:
{text}
"""
    return f"""
Extrais uniquement les informations de facture.
Retourne un JSON conforme au schéma.

Règles:
- type_document = facture
- Si une valeur est absente, retourne null
- numero_facture = numéro principal de la facture
- date_emission et date_paiement_prevue entre 2020 et 2100 sinon null
- si la date de paiement n'est pas explicitement présente, utilise les conditions de paiement
- conditions_paiement = libellé exact des conditions de règlement si présent
- si des bons de livraison sont référencés, renseigne bons_livraisons avec leurs numéros
- n'invente jamais une année différente de celle écrite dans le document
- prix_HT_5_5pct, prix_HT_10pct, prix_HT_20pct = montants HT totaux de la facture par taux de TVA
- normalise les montants en nombres décimaux
- nom_fournisseur doit être exactement une des valeurs suivantes : {fournisseurs_str}

Texte:
{text}
"""

def link_documents(factures: list[dict], bons: list[dict]) -> tuple[list[dict], list[dict]]:
    bl_to_facture = {}
    for f in factures:
        for bl_num in f.get("bons_livraisons", []):
            bl_to_facture[str(bl_num).strip()] = f.get("numero_facture")

    for bon in bons:
        bon_num = bon.get("numero_bon_livraison")
        if bon_num:
            bon["numero_facture_rattachee"] = bl_to_facture.get(str(bon_num).strip())

    return factures, bons

def fetch_files_from_api():
    os.makedirs(DOSSIER_FACTURES, exist_ok=True)

# ---------------------------------------------------------------------------
# Mapping nom fournisseur app → nom affiché dans Achats Cons
# ---------------------------------------------------------------------------
FOURNISSEUR_DISPLAY = {
    "SYSCO":     "Sysco",
    "AMBELYS":   "Ambelys",
    "TERREAZUR": "TerreAzur",
}

def write_to_achats_cons(
    factures: list[dict],
    bons: list[dict],
    template_path: str,
    output_path: str,
    fournisseur_display: dict[str, str] | None = None,
) -> int:
    """
    Ouvre le fichier de suivi trésorerie MLC, efface les lignes gérées par
    l'appli (col C = nos fournisseurs), puis réinsère selon la structure réelle :

      - UNE LIGNE PAR BL : col E = N° BL, col F = date du BL,
        montants HT portés par le BL (prix_HT_5_5pct / 10pct / 20pct du BL)
      - Facture sans BL : une seule ligne, col E vide, col F = date_emission,
        montants HT de la facture

    Colonnes saisies : C (Fournisseur), D (Fact), E (BL), F (Date),
                       I (HT 5.5), J (HT 10), K (HT 20), S (Date paiement), W (Commentaires)
    Colonnes formule : A, B, G, H, L, M, N, O, P, Q, R, T, U, V, X, Y

    Retourne le nombre de lignes insérées.
    """
    import openpyxl
    from openpyxl import load_workbook
    from datetime import date as _date

    _display = fournisseur_display if fournisseur_display is not None else FOURNISSEUR_DISPLAY
    MLC_FOURNISSEURS = {v.lower() for v in _display.values()}

    # Index facture → liste de dicts BL complets (avec leurs montants propres)
    bl_par_facture: dict[str, list[dict]] = {}
    for bon in bons:
        fac_num = bon.get("numero_facture_rattachee")
        bl_num  = bon.get("numero_bon_livraison")
        if fac_num and bl_num:
            bl_par_facture.setdefault(str(fac_num), [])
            if not any(b["numero_bon_livraison"] == bl_num
                       for b in bl_par_facture[str(fac_num)]):
                bl_par_facture[str(fac_num)].append(bon)

    wb = load_workbook(template_path, keep_vba=True)
    ws = wb["Achats Cons"]

    # 1. Effacer les lignes MLC existantes
    mlc_rows: list[int] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        cell_c = row[2]
        if cell_c and str(cell_c).strip().lower() in MLC_FOURNISSEURS:
            mlc_rows.append(row_idx)
    for r in mlc_rows:
        for c in range(1, 26):
            ws.cell(r, c).value = None

    # 2. Première ligne vide disponible
    first_empty = 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            first_empty += 1
        else:
            break

    # Helpers
    def _to_date(v):
        if v is None:
            return None
        if isinstance(v, str):
            try:
                return _date.fromisoformat(v)
            except ValueError:
                return None
        if hasattr(v, "date"):
            return v.date()
        if isinstance(v, _date):
            return v
        return None

    def _to_float(v):
        if v is None:
            return None
        try:
            f = float(v)
            return f if f != 0.0 else None
        except (ValueError, TypeError):
            return None

    def _write_row(r, fournisseur, num_facture, num_bl,
                   date_f, ht_55, ht_10, ht_20, date_paiement, commentaire):
        ws.cell(r, 1).value  = f'=IF(AND(B{r}>=TDB!$B$6,B{r}<=TDB!$D$6),"Oui","")'
        ws.cell(r, 2).value  = f'=IF(G{r}<10,H{r}&0&G{r},H{r}&G{r})'
        ws.cell(r, 7).value  = f'=IF(F{r}="","",MONTH(F{r}))'
        ws.cell(r, 8).value  = f'=IF(F{r}="","",YEAR(F{r}))'
        ws.cell(r, 12).value = f'=IF(AND(I{r}="",J{r}="",K{r}=""),"",SUM(I{r}:K{r}))'
        ws.cell(r, 13).value = f'=IF(I{r}="","",I{r}*0.055)'
        ws.cell(r, 14).value = f'=IF(J{r}="","",J{r}*0.1)'
        ws.cell(r, 15).value = f'=IF(K{r}="","",K{r}*0.2)'
        ws.cell(r, 16).value = f'=IF(AND(M{r}="",N{r}="",O{r}=""),"",SUM(M{r}:O{r}))'
        ws.cell(r, 17).value = f'=IF(AND(L{r}="",P{r}=""),"",L{r}+P{r})'
        ws.cell(r, 18).value = f'=IFERROR(INDEX(Inputs!$C:$C,MATCH(C{r},Inputs!$B:$B,0)),"")'
        ws.cell(r, 20).value = f'=IF(I{r}="","",IF(M{r}=0,"",IF(ROUND(M{r}/I{r},3)=0.055,"OK","Erreur")))'
        ws.cell(r, 21).value = f'=IF(J{r}="","",IF(N{r}=0,"",IF(ROUND(N{r}/J{r},3)=0.1,"OK","Erreur")))'
        ws.cell(r, 22).value = f'=IF(K{r}="","",IF(O{r}=0,"",IF(ROUND(O{r}/K{r},3)=0.2,"OK","Erreur")))'
        ws.cell(r, 24).value = f'=S{r}&"-"&C{r}&"-"&TEXT(Q{r},"0.00")'
        ws.cell(r, 25).value = f'=IFERROR(INDEX(Inputs!$D:$D,MATCH(\'Achats Cons\'!C{r},Inputs!$B:$B,0)),"")'
        ws.cell(r, 3).value  = fournisseur
        ws.cell(r, 4).value  = num_facture
        ws.cell(r, 5).value  = num_bl or None
        ws.cell(r, 6).value  = date_f
        ws.cell(r, 9).value  = ht_55
        ws.cell(r, 10).value = ht_10
        ws.cell(r, 11).value = ht_20
        ws.cell(r, 19).value = date_paiement
        ws.cell(r, 23).value = commentaire or None
        if date_f:
            ws.cell(r, 6).number_format = "DD/MM/YYYY"
        if date_paiement:
            ws.cell(r, 19).number_format = "DD/MM/YYYY"

    # 3. Insérer : une ligne par BL, ou une ligne par facture sans BL
    inserted = 0
    inserted_bl_nums: set[str] = set()
    factures_ids = {
        str(f.get("numero_facture"))
        for f in factures
        if f.get("numero_facture")
    }

    for facture in factures:
        fournisseur_raw = facture.get("nom_fournisseur") or ""
        fournisseur     = _display.get(fournisseur_raw.upper(), fournisseur_raw)
        num_facture     = facture.get("numero_facture")
        date_emission   = _to_date(facture.get("date_emission"))
        date_paiement   = _to_date(facture.get("date_paiement_prevue"))
        commentaire     = facture.get("fichier_source") or facture.get("fichier_stocke") or ""
        bls             = bl_par_facture.get(str(num_facture), []) if num_facture else []

        if bls:
            # Une ligne par BL — montants portés par le BL
            for bon in bls:
                num_bl  = bon.get("numero_bon_livraison")
                date_bl = _to_date(bon.get("date_livraison")) or date_emission
                ht_55   = _to_float(bon.get("prix_HT_5_5pct"))
                ht_10   = _to_float(bon.get("prix_HT_10pct"))
                ht_20   = _to_float(bon.get("prix_HT_20pct"))
                _write_row(first_empty + inserted, fournisseur, num_facture, num_bl,
                           date_bl, ht_55, ht_10, ht_20, date_paiement, commentaire)
                inserted += 1
                if num_bl:
                    inserted_bl_nums.add(str(num_bl))
        else:
            # Pas de BL → une seule ligne avec les montants de la facture
            ht_55 = _to_float(facture.get("prix_HT_5_5pct"))
            ht_10 = _to_float(facture.get("prix_HT_10pct"))
            ht_20 = _to_float(facture.get("prix_HT_20pct"))
            _write_row(first_empty + inserted, fournisseur, num_facture, None,
                       date_emission, ht_55, ht_10, ht_20, date_paiement, commentaire)
            inserted += 1

    # 4. Ajouter les BL sans facture (ou rattachés à une facture absente)
    for bon in bons:
        num_bl = bon.get("numero_bon_livraison")
        if not num_bl:
            continue
        fac_num = bon.get("numero_facture_rattachee")
        if str(num_bl) in inserted_bl_nums:
            continue
        if fac_num and str(fac_num) in factures_ids:
            continue

        fournisseur_raw = bon.get("nom_fournisseur") or ""
        fournisseur = _display.get(str(fournisseur_raw).upper(), fournisseur_raw)
        date_bl = _to_date(bon.get("date_livraison"))
        ht_55 = _to_float(bon.get("prix_HT_5_5pct"))
        ht_10 = _to_float(bon.get("prix_HT_10pct"))
        ht_20 = _to_float(bon.get("prix_HT_20pct"))
        commentaire = bon.get("fichier_source") or bon.get("fichier_stocke") or ""

        _write_row(
            first_empty + inserted,
            fournisseur,
            None,
            num_bl,
            date_bl,
            ht_55,
            ht_10,
            ht_20,
            None,
            commentaire,
        )
        inserted += 1

    atomic_save_workbook(wb, output_path)
    wb.close()
    return inserted


if __name__ == "__main__":
    print("Démarrage du traitement des documents...")
    fetch_files_from_api()
    factures_df, bons_df = process_all_documents()
    print(f"Traitement terminé. {len(factures_df)} factures et {len(bons_df)} bons de livraison extraits.")