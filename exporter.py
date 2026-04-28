"""
Module d'export : génère le fichier XLSM depuis les données SQLite.

Approche ZIP chirurgicale :
- Le XLSM est un fichier ZIP contenant un XML par onglet.
- On ne charge JAMAIS le workbook complet avec openpyxl (trop lent : 42 onglets, ~200 Mo de XML).
- On génère chaque onglet cible dans un workbook openpyxl VIDE (léger),
  on extrait son XML, et on le réinjecte directement dans le ZIP du fichier cible.
- Les 38 autres onglets (TCD, P&L, analyses…) ne sont jamais touchés.

Onglets gérés :
  - Achats Cons   (sheet16.xml) : factures + BL
  - Autres achats (sheet32.xml) : autres achats
  - DOMINO        (sheet10.xml) : données journalières
  - Inputs        (sheet9.xml)  : fournisseurs
"""

from __future__ import annotations

import io
import os
import shutil
import tempfile
import zipfile
from datetime import date as dateclass
from xml.etree import ElementTree as ET

import openpyxl
import repositories as repo
import domino
from xlsm_safe import is_valid_xlsm

CURRENT_XLSM_PATH = "output/Suivi trésorerie MLC.xlsm"

# Mapping nom d'onglet → chemin dans le ZIP (stable pour ce fichier)
_SHEET_ZIP_PATHS: dict[str, str] = {
    "Inputs":        "xl/worksheets/sheet9.xml",
    "DOMINO":        "xl/worksheets/sheet10.xml",
    "Achats Cons":   "xl/worksheets/sheet16.xml",
    "Autres achats": "xl/worksheets/sheet32.xml",
}


# ---------------------------------------------------------------------------
# Résolution dynamique des chemins ZIP (robustesse si le fichier change)
# ---------------------------------------------------------------------------

def _resolve_sheet_zip_paths(xlsm_path: str) -> dict[str, str]:
    """
    Lit workbook.xml + workbook.xml.rels pour mapper nom d'onglet → chemin ZIP.
    Retourne le mapping pour les 4 onglets cibles uniquement.
    Utilise le mapping statique _SHEET_ZIP_PATHS comme fallback.
    """
    targets = set(_SHEET_ZIP_PATHS.keys())
    result: dict[str, str] = {}
    try:
        with zipfile.ZipFile(xlsm_path, "r") as zf:
            wb_xml = zf.read("xl/workbook.xml")
            rels_xml = zf.read("xl/_rels/workbook.xml.rels")

        ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        ns_r    = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

        root = ET.fromstring(wb_xml)
        name_to_rid: dict[str, str] = {}
        for sheet in root.findall(f".//{{{ns_main}}}sheet"):
            name = sheet.get("name", "")
            if name in targets:
                name_to_rid[name] = sheet.get(f"{{{ns_r}}}id", "")

        rels_root = ET.fromstring(rels_xml)
        rid_to_target: dict[str, str] = {}
        for rel in rels_root:
            rid_to_target[rel.get("Id", "")] = rel.get("Target", "")

        for name, rid in name_to_rid.items():
            target = rid_to_target.get(rid, "")
            if target:
                if target.startswith("/"):
                    result[name] = target.lstrip("/")
                else:
                    result[name] = f"xl/{target}" if not target.startswith("xl/") else target
    except Exception as e:
        print(f"[WARN] _resolve_sheet_zip_paths: {e} — utilisation du mapping statique")

    for name, path in _SHEET_ZIP_PATHS.items():
        if name not in result:
            result[name] = path

    return result


# ---------------------------------------------------------------------------
# Extraction des lignes de header depuis le template ZIP
# ---------------------------------------------------------------------------

def _extract_header_rows_xml(xlsm_path: str, sheet_zip_path: str, nb_header_rows: int) -> list[bytes]:
    """
    Lit les `nb_header_rows` premières lignes de l'onglet dans le ZIP du template
    et les retourne comme une liste de bytes XML bruts (un élément <row> par entrée).
    Ces lignes seront préfixées dans le XML généré pour préserver les headers.
    """
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    try:
        with zipfile.ZipFile(xlsm_path, "r") as zf:
            xml = zf.read(sheet_zip_path)
        root = ET.fromstring(xml)
        rows = root.findall(f"{{{ns}}}sheetData/{{{ns}}}row")
        header_rows = rows[:nb_header_rows]
        return [ET.tostring(r, encoding="unicode").encode("utf-8") for r in header_rows]
    except Exception as e:
        print(f"[WARN] _extract_header_rows_xml({sheet_zip_path}): {e}")
        return []


# ---------------------------------------------------------------------------
# Génération XML d'un onglet via un workbook openpyxl vide (léger)
# ---------------------------------------------------------------------------

def _build_sheet_xml(populate_fn, header_rows_xml: list[bytes] | None = None) -> bytes:
    """
    Crée un workbook openpyxl vide, appelle populate_fn(ws) pour remplir
    la feuille active (les données commencent toujours à la ligne 1 dans openpyxl),
    puis extrait le XML.

    Si `header_rows_xml` est fourni :
    - Les lignes de données sont renumérotées (r, ref cellules) en les décalant
      de len(header_rows_xml) lignes vers le bas.
    - Les lignes de header sont insérées en tête du sheetData.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    populate_fn(ws)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)

    raw_xml = zipfile.ZipFile(buf, "r").read("xl/worksheets/sheet1.xml")

    if not header_rows_xml:
        return raw_xml

    ns_str = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    offset = len(header_rows_xml)

    root = ET.fromstring(raw_xml)
    sheet_data = root.find(f"{{{ns_str}}}sheetData")
    if sheet_data is None:
        return raw_xml

    # Renuméroter toutes les lignes de données (r + références cellules)
    import re as _re
    for row_elem in sheet_data.findall(f"{{{ns_str}}}row"):
        old_r = int(row_elem.get("r", "1"))
        new_r = old_r + offset
        row_elem.set("r", str(new_r))
        for cell in row_elem.findall(f"{{{ns_str}}}c"):
            ref = cell.get("r", "")
            # Ref format : lettre(s) + chiffre(s), ex: "A1", "BC12"
            new_ref = _re.sub(r"(\d+)$", str(new_r), ref)
            cell.set("r", new_ref)

    # Insérer les lignes header au début du sheetData
    for row_xml in reversed(header_rows_xml):
        row_elem = ET.fromstring(row_xml.decode("utf-8"))
        sheet_data.insert(0, row_elem)

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


# ---------------------------------------------------------------------------
# Fonctions de remplissage par onglet
# ---------------------------------------------------------------------------

def _populate_inputs(ws) -> int:
    """Données fournisseurs à partir de la ligne 1 (le header sera injecté avant via décalage)."""
    fournisseurs = repo.list_fournisseurs()
    for idx, fourn in enumerate(fournisseurs):
        r = 1 + idx  # commence à 1, sera décalé à 2 après injection du header
        ws.cell(r, 2).value = fourn.get("nom_affiche")
        ws.cell(r, 3).value = fourn.get("conditions_paiement")
        ws.cell(r, 4).value = fourn.get("categorie")
        ws.cell(r, 5).value = fourn.get("mode_paiement")
        ws.cell(r, 6).value = fourn.get("frequence")
        ws.cell(r, 7).value = fourn.get("mois")
    return len(fournisseurs)


def _populate_autres_achats(ws) -> int:
    """Données autres achats à partir de la ligne 1 (le header sera injecté avant via décalage)."""
    autres_achats = repo.list_autres_achats()
    col_mapping = {
        "fournisseur": 1, "categorie": 2, "num_facture": 3, "num_bl": 4,
        "date": 5, "ht_0": 8, "ht_2_1": 9, "ht_5_5": 10, "ht_10": 11,
        "ht_20": 12, "conditions": 21, "date_paiement": 22,
        "amortissable": 26, "ref_denotage": 27,
    }
    for idx, achat in enumerate(autres_achats):
        r = 1 + idx  # commence à 1, sera décalé à 2 après injection du header
        for field, col in col_mapping.items():
            v = achat.get(field)
            if v is not None:
                ws.cell(r, col).value = v
    return len(autres_achats)


def _populate_achats_cons(ws, factures: list[dict], bons: list[dict], fournisseur_display: dict[str, str]) -> int:
    """
    Écrit les factures + BL dans ws (onglet Achats Cons vide).
    Reproduit la logique de write_to_achats_cons sans charger le XLSM complet.
    """
    from datetime import date as _date

    MLC_FOURNISSEURS = {v.lower() for v in fournisseur_display.values()}

    bl_par_facture: dict[str, list[dict]] = {}
    for bon in bons:
        fac_num = bon.get("numero_facture_rattachee")
        bl_num  = bon.get("numero_bon_livraison")
        if fac_num and bl_num:
            bl_par_facture.setdefault(str(fac_num), [])
            if not any(b["numero_bon_livraison"] == bl_num for b in bl_par_facture[str(fac_num)]):
                bl_par_facture[str(fac_num)].append(bon)

    def _to_date(v):
        if v is None: return None
        if isinstance(v, str):
            try: return _date.fromisoformat(v)
            except ValueError: return None
        if hasattr(v, "date"): return v.date()
        if isinstance(v, _date): return v
        return None

    def _to_float(v):
        if v is None: return None
        try:
            f = float(v)
            return f if f != 0.0 else None
        except (ValueError, TypeError):
            return None

    def _write_row(r, fournisseur, num_facture, num_bl, date_f, ht_55, ht_10, ht_20, date_paiement, commentaire):
        ws.cell(r, 1).value  = f'=IF(AND(B{r}>=TDB!$B$6,B{r}<=TDB!$D$6),"Oui","")'
        ws.cell(r, 2).value  = f'=IF(G{r}<10,H{r}&0&G{r},H{r}&G{r})'
        ws.cell(r, 7).value  = f'=IF(F{r}="","",MONTH(F{r}))'
        ws.cell(r, 8).value  = f'=IF(F{r}="","",YEAR(F{r}))'
        ws.cell(r, 12).value = f'=IF(AND(I{r}="",J{r}="",K{r}=""),"",SUM(I{r}:K{r}))'
        ws.cell(r, 13).value = f'=IF(I{r}="","",I{r}*0.055)'
        ws.cell(r, 14).value = f'=IF(J{r}="","",J{r}*0.1)'
        ws.cell(r, 15).value = f'=IF(K{r}="","",K{r}*0.2)'
        ws.cell(r, 16).value = f'=IF(AND(M{r}="",N{r}="",O{r}=""),"",SUM(M{r}:O{r}))'
        ws.cell(r, 17).value = f'=IF(AND(L{r}="",P{r}=""),"",L{r}+P{r})'
        ws.cell(r, 18).value = f'=IFERROR(INDEX(Inputs!$C:$C,MATCH(C{r},Inputs!$B:$B,0)),"")'
        ws.cell(r, 20).value = f'=IF(I{r}="","",IF(M{r}=0,"",IF(ROUND(M{r}/I{r},3)=0.055,"OK","Erreur")))'
        ws.cell(r, 21).value = f'=IF(J{r}="","",IF(N{r}=0,"",IF(ROUND(N{r}/J{r},3)=0.1,"OK","Erreur")))'
        ws.cell(r, 22).value = f'=IF(K{r}="","",IF(O{r}=0,"",IF(ROUND(O{r}/K{r},3)=0.2,"OK","Erreur")))'
        ws.cell(r, 24).value = f'=S{r}&"-"&C{r}&"-"&TEXT(Q{r},"0.00")'
        ws.cell(r, 25).value = f'=IFERROR(INDEX(Inputs!$D:$D,MATCH(\'Achats Cons\'!C{r},Inputs!$B:$B,0)),"")'
        ws.cell(r, 3).value  = fournisseur
        ws.cell(r, 4).value  = num_facture
        ws.cell(r, 5).value  = num_bl or None
        ws.cell(r, 6).value  = date_f
        ws.cell(r, 9).value  = ht_55
        ws.cell(r, 10).value = ht_10
        ws.cell(r, 11).value = ht_20
        ws.cell(r, 19).value = date_paiement
        ws.cell(r, 23).value = commentaire or None
        if date_f:
            ws.cell(r, 6).number_format = "DD/MM/YYYY"
        if date_paiement:
            ws.cell(r, 19).number_format = "DD/MM/YYYY"

    inserted = 0
    row = 1  # commence à 1, sera décalé à 2 après injection du header
    inserted_bl_nums: set[str] = set()
    factures_ids = {str(f.get("numero_facture")) for f in factures if f.get("numero_facture")}

    for facture in factures:
        fournisseur_raw = facture.get("nom_fournisseur") or ""
        fournisseur     = fournisseur_display.get(fournisseur_raw.upper(), fournisseur_raw)
        num_facture     = facture.get("numero_facture")
        date_emission   = _to_date(facture.get("date_emission"))
        date_paiement   = _to_date(facture.get("date_paiement_prevue"))
        commentaire     = facture.get("fichier_source") or facture.get("fichier_stocke") or ""
        bls             = bl_par_facture.get(str(num_facture), []) if num_facture else []

        if bls:
            for bon in bls:
                num_bl  = bon.get("numero_bon_livraison")
                date_bl = _to_date(bon.get("date_livraison")) or date_emission
                _write_row(row, fournisseur, num_facture, num_bl,
                           date_bl, _to_float(bon.get("prix_HT_5_5pct")),
                           _to_float(bon.get("prix_HT_10pct")), _to_float(bon.get("prix_HT_20pct")),
                           date_paiement, commentaire)
                row += 1; inserted += 1
                if num_bl: inserted_bl_nums.add(str(num_bl))
        else:
            _write_row(row, fournisseur, num_facture, None, date_emission,
                       _to_float(facture.get("prix_HT_5_5pct")), _to_float(facture.get("prix_HT_10pct")),
                       _to_float(facture.get("prix_HT_20pct")), date_paiement, commentaire)
            row += 1; inserted += 1

    for bon in bons:
        num_bl = bon.get("numero_bon_livraison")
        if not num_bl: continue
        if str(num_bl) in inserted_bl_nums: continue
        fac_num = bon.get("numero_facture_rattachee")
        if fac_num and str(fac_num) in factures_ids: continue
        fournisseur_raw = bon.get("nom_fournisseur") or ""
        fournisseur = fournisseur_display.get(str(fournisseur_raw).upper(), fournisseur_raw)
        _write_row(row, fournisseur, None, num_bl, _to_date(bon.get("date_livraison")),
                   _to_float(bon.get("prix_HT_5_5pct")), _to_float(bon.get("prix_HT_10pct")),
                   _to_float(bon.get("prix_HT_20pct")), None,
                   bon.get("fichier_source") or bon.get("fichier_stocke") or "")
        row += 1; inserted += 1

    return inserted


def _populate_domino(ws) -> tuple[int, list[str]]:
    jours = repo.list_domino_jours()
    errors: list[str] = []
    inserted = 0
    for jour_row in jours:
        try:
            d = jour_row.get("date")
            parsed_date = dateclass.fromisoformat(d) if isinstance(d, str) else d
            domino_data = domino.DominoJourData(
                date=parsed_date,
                filename=jour_row.get("filename", "export"),
                ca_ttc_matin=float(jour_row.get("ca_ttc_matin") or 0),
                ca_ttc_midi=float(jour_row.get("ca_ttc_midi") or 0),
                ca_ttc_apm=float(jour_row.get("ca_ttc_apm") or 0),
                ca_ttc_soir=float(jour_row.get("ca_ttc_soir") or 0),
                ca_ttc_uber=float(jour_row.get("ca_ttc_uber") or 0),
                ca_ttc_deliveroo=float(jour_row.get("ca_ttc_deliveroo") or 0),
                ca_ttc_total=float(jour_row.get("ca_ttc_total") or 0),
                tva_total=float(jour_row.get("tva_total") or 0),
                tva_55=float(jour_row.get("tva_55") or 0),
                tva_10=float(jour_row.get("tva_10") or 0),
                especes=float(jour_row.get("especes") or 0),
                carte_bancaire=float(jour_row.get("carte_bancaire") or 0),
                cb_link=float(jour_row.get("cb_link") or 0),
                belorder=float(jour_row.get("belorder") or 0),
                uber_eats=float(jour_row.get("uber_eats") or 0),
                deliveroo_paiement=float(jour_row.get("deliveroo_paiement") or 0),
                total_encaissements=float(jour_row.get("total_encaissements") or 0),
                nb_clients_matin=int(jour_row.get("nb_clients_matin") or 0),
                nb_clients_midi=int(jour_row.get("nb_clients_midi") or 0),
                nb_clients_soir=int(jour_row.get("nb_clients_soir") or 0),
                total_clients=int(jour_row.get("total_clients") or 0),
            )
            result = domino._write_data_on_open_sheet(ws, domino_data, overwrite=False)
            if not result.get("skipped"):
                inserted += 1
        except Exception as e:
            errors.append(f"DOMINO {jour_row.get('date')}: {e}")
    return inserted, errors


def export_to_xlsm(output_path: str = CURRENT_XLSM_PATH) -> dict:
    """
    Injecte les données SQLite dans les 4 onglets cibles du XLSM
    en remplaçant chirurgicalement leur XML dans le ZIP.

    Les 38 autres onglets (TCD, P&L, analyses…) ne sont jamais chargés
    ni modifiés → temps d'exécution < 10 s au lieu de 3 min.
    """
    if not is_valid_xlsm(output_path):
        raise FileNotFoundError(f"Fichier XLSM invalide ou introuvable : '{output_path}'")

    sheet_paths = _resolve_sheet_zip_paths(output_path)
    stats: dict = {"domino_jours": 0, "autres_achats_lignes": 0,
                   "inputs_fournisseurs": 0, "achats_cons_lignes": 0, "errors": []}

    # Source des headers : toujours le template original (jamais modifié par nos exports)
    # Si le template n'existe pas, on tombe back sur output_path lui-même.
    template_for_headers = "output/template.xlsm" if os.path.exists("output/template.xlsm") else output_path
    template_sheet_paths = _resolve_sheet_zip_paths(template_for_headers)

    # --- 1. Achats Cons : 1 ligne de header (ligne 1 = noms de colonnes) ---
    try:
        factures = repo.list_factures()
        bons = repo.list_bons()
        fournisseur_display = repo.fournisseur_display_map()
        header_achats = _extract_header_rows_xml(template_for_headers, template_sheet_paths["Achats Cons"], 1)
        nb_achats = [0]
        def _fill_achats(ws):
            nb_achats[0] = _populate_achats_cons(ws, factures, bons, fournisseur_display)
        new_xmls_achats: dict[str, bytes] = {
            "Achats Cons": _build_sheet_xml(_fill_achats, header_achats)
        }
        _inject_xmls_into_zip(output_path, sheet_paths, new_xmls_achats)
        stats["achats_cons_lignes"] = nb_achats[0]
    except Exception as e:
        stats["errors"].append(f"Achats Cons: {e}")
        print(f"[WARN] Achats Cons: {e}")

    # --- 2. Générer les XML des 3 autres onglets dans des workbooks vides ---
    new_xmls: dict[str, bytes] = {}

    # Inputs : 1 ligne de header (ligne 1 = titre de section)
    try:
        header_inputs = _extract_header_rows_xml(template_for_headers, template_sheet_paths["Inputs"], 1)
        nb_inputs = [0]
        def _fill_inputs(ws):
            nb_inputs[0] = _populate_inputs(ws)
        new_xmls["Inputs"] = _build_sheet_xml(_fill_inputs, header_inputs)
        stats["inputs_fournisseurs"] = nb_inputs[0]
    except Exception as e:
        stats["errors"].append(f"Inputs: {e}")
        print(f"[WARN] Inputs: {e}")

    # Autres achats : 1 ligne de header (ligne 1 = noms de colonnes)
    try:
        header_autres = _extract_header_rows_xml(template_for_headers, template_sheet_paths["Autres achats"], 1)
        nb_autres = [0]
        def _fill_autres(ws):
            nb_autres[0] = _populate_autres_achats(ws)
        new_xmls["Autres achats"] = _build_sheet_xml(_fill_autres, header_autres)
        stats["autres_achats_lignes"] = nb_autres[0]
    except Exception as e:
        stats["errors"].append(f"Autres achats: {e}")
        print(f"[WARN] Autres achats: {e}")

    # DOMINO : 1 ligne de header (ligne 1)
    try:
        header_domino = _extract_header_rows_xml(template_for_headers, template_sheet_paths["DOMINO"], 1)
        domino_results = [0, []]
        def _fill_domino(ws):
            nb, errs = _populate_domino(ws)
            domino_results[0] = nb
            domino_results[1] = errs
        new_xmls["DOMINO"] = _build_sheet_xml(_fill_domino, header_domino)
        stats["domino_jours"] = domino_results[0]
        stats["errors"].extend(domino_results[1])
    except Exception as e:
        stats["errors"].append(f"DOMINO: {e}")
        print(f"[WARN] DOMINO: {e}")

    # --- 3. Réinjection ZIP chirurgicale ---
    if new_xmls:
        _inject_xmls_into_zip(output_path, sheet_paths, new_xmls)

    return {"status": "success", "output_file": output_path, **stats}


def _inject_xmls_into_zip(
    xlsm_path: str,
    sheet_paths: dict[str, str],
    new_xmls: dict[str, bytes],
) -> None:
    """
    Remplace les XML des onglets cibles dans le ZIP du XLSM.
    Écrit dans un fichier temporaire puis remplace atomiquement.
    Crée un backup .lastgood.bak avant toute modification.
    """
    backup_path = f"{xlsm_path}.lastgood.bak"
    if is_valid_xlsm(xlsm_path):
        shutil.copy2(xlsm_path, backup_path)

    target_dir = os.path.dirname(os.path.abspath(xlsm_path)) or "."
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsm", prefix="export_zip_", dir=target_dir)
    os.close(fd)

    try:
        with zipfile.ZipFile(xlsm_path, "r") as zin, \
             zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:

            to_replace = {
                sheet_paths[name]: xml_bytes
                for name, xml_bytes in new_xmls.items()
                if name in sheet_paths
            }

            for item in zin.infolist():
                if item.filename in to_replace:
                    zout.writestr(item, to_replace[item.filename])
                else:
                    zout.writestr(item, zin.read(item.filename))

        if not is_valid_xlsm(tmp_path):
            raise RuntimeError("Le fichier ZIP généré est invalide.")

        os.replace(tmp_path, xlsm_path)

    except Exception:
        if os.path.exists(backup_path) and not is_valid_xlsm(xlsm_path):
            os.replace(backup_path, xlsm_path)
        raise
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


if __name__ == "__main__":
    import time
    t = time.time()
    result = export_to_xlsm()
    print(f"Export en {round(time.time()-t, 2)}s :", result)
