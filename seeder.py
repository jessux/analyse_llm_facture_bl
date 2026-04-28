"""
Seeder : remplit la BDD SQLite à partir du XLSM courant et du JSON DOMINO.

Appelé au démarrage si la BDD est vide.
"""

from __future__ import annotations

import json
import os
from datetime import date, datetime
from typing import Any

import openpyxl

import db
import repositories as repo

DEFAULT_XLSM_CANDIDATES = (
    "output/Suivi trésorerie MLC.xlsm",
    "output/Suivi trésorerie MLC - Copie.xlsm",
)
DEFAULT_DOMINO_JSON = "output/domino_imports.json"

# 3 fournisseurs historiques (créés systématiquement avec patterns connus)
HISTORIC_FOURNISSEURS: dict[str, dict[str, Any]] = {
    "SYSCO":     {"nom_affiche": "Sysco",     "patterns": ["sysco"]},
    "AMBELYS":   {"nom_affiche": "Ambelys",   "patterns": ["ambelys"]},
    "TERREAZUR": {"nom_affiche": "TerreAzur", "patterns": ["terreazur", "terre azur"]},
}


def _xlsm_path() -> str | None:
    for candidate in DEFAULT_XLSM_CANDIDATES:
        if os.path.exists(candidate):
            return candidate
    return None


def _to_iso_date(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        # Laisse repositories normaliser
        return v
    return None


def _to_str_or_none(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_float_or_none(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Seed Inputs (fournisseurs)
# ---------------------------------------------------------------------------

def seed_fournisseurs(xlsm_path: str | None) -> int:
    """
    Crée d'abord les fournisseurs historiques, puis ajoute ceux trouvés dans
    l'onglet Inputs (col B = nom). Conserve les éventuelles patterns existantes.
    """
    # Historiques
    for fid, conf in HISTORIC_FOURNISSEURS.items():
        if repo.get_fournisseur(fid) is None:
            repo.upsert_fournisseur(
                id=fid,
                nom_affiche=conf["nom_affiche"],
                patterns=conf["patterns"],
            )

    if not xlsm_path or not os.path.exists(xlsm_path):
        return len(HISTORIC_FOURNISSEURS)

    try:
        wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[SEED] Impossible d'ouvrir {xlsm_path} pour Inputs: {e}")
        return len(HISTORIC_FOURNISSEURS)

    if "Inputs" not in wb.sheetnames:
        wb.close()
        return len(HISTORIC_FOURNISSEURS)

    ws = wb["Inputs"]
    inserted = 0
    try:
        # Localiser le début et la fin de la liste fournisseurs (col B).
        # La table commence à la 1ère ligne après "Liste des fournisseurs marchandises"
        # et se termine à la 1ère ligne dont col B est vide.
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        start_row = None
        for i, r in enumerate(rows):
            if len(r) > 1 and isinstance(r[1], str) and "liste des fournisseurs" in r[1].lower():
                start_row = i + 1
                break

        if start_row is None:
            # Fallback : la convention historique place la liste en B3..
            start_row = 2

        for r in rows[start_row:]:
            if not r or len(r) < 7:
                break
            nom = _to_str_or_none(r[1])  # col B
            if not nom:
                break  # fin de la table
            # Filtre robustesse : un fournisseur valide n'est pas purement numérique
            # ni trop court (cas où on déborde sur une autre table latérale).
            if nom.isdigit() or len(nom) < 2:
                break

            fid = repo.make_supplier_key(nom)
            existing = repo.get_fournisseur(fid)

            conditions = _to_str_or_none(r[2])     # col C
            categorie = _to_str_or_none(r[3])      # col D
            mode_paiement = _to_str_or_none(r[4])  # col E
            frequence = _to_str_or_none(r[5])      # col F
            mois = _to_str_or_none(r[6])           # col G

            patterns = (existing or {}).get("patterns") or [nom.lower()]
            display = (existing or {}).get("nom_affiche") or nom

            repo.upsert_fournisseur(
                id=fid,
                nom_affiche=display,
                patterns=patterns,
                conditions_paiement=conditions,
                categorie=categorie,
                mode_paiement=mode_paiement,
                frequence=frequence,
                mois=mois,
            )
            if not existing:
                inserted += 1
    finally:
        wb.close()

    return inserted + len(HISTORIC_FOURNISSEURS)


# ---------------------------------------------------------------------------
# Seed Achats Cons (factures + BL)
# ---------------------------------------------------------------------------

def seed_achats_cons(xlsm_path: str) -> tuple[int, int]:
    """
    Reproduit la logique de `_startup_load_excel` mais écrit en BDD.
    Retourne (nb_factures, nb_bons).
    """
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)
    if "Achats Cons" not in wb.sheetnames:
        wb.close()
        return (0, 0)
    ws = wb["Achats Cons"]

    # Index display_name(lower) -> id depuis la BDD
    display_to_key = {
        f["nom_affiche"].lower(): f["id"] for f in repo.list_fournisseurs()
    }

    factures: dict[str, dict] = {}   # numero_facture -> dict (HT agrégés)
    bons: dict[str, dict] = {}       # numero_bl -> dict
    fac_to_bls: dict[str, list[str]] = {}

    storage_dir = "storage"

    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            fournisseur_raw = row[2]
            if not fournisseur_raw:
                continue
            fournisseur_display = str(fournisseur_raw).strip()
            fkey = display_to_key.get(fournisseur_display.lower())
            if not fkey:
                # Auto-création (cas robustesse)
                fkey = repo.ensure_fournisseur_from_display(fournisseur_display)
                display_to_key[fournisseur_display.lower()] = fkey

            num_facture = _to_str_or_none(row[3])
            num_bl = _to_str_or_none(row[4])
            date_emission = _to_iso_date(row[5])
            ht_55 = _to_float_or_none(row[8])
            ht_10 = _to_float_or_none(row[9])
            ht_20 = _to_float_or_none(row[10])
            date_paiement = _to_iso_date(row[18])
            commentaire = _to_str_or_none(row[22])

            if not num_facture and not num_bl:
                continue

            # Facture (création / agrégation HT multi-lignes)
            if num_facture:
                rec = factures.get(num_facture)
                if rec is None:
                    fichier_stocke = None
                    if commentaire and os.path.exists(os.path.join(storage_dir, commentaire)):
                        fichier_stocke = commentaire
                    rec = {
                        "numero_facture": num_facture,
                        "nom_fournisseur": fkey,
                        "date_emission": date_emission,
                        "date_paiement_prevue": date_paiement,
                        "prix_HT_5_5pct": ht_55,
                        "prix_HT_10pct": ht_10,
                        "prix_HT_20pct": ht_20,
                        "fichier_source": commentaire or "",
                        "fichier_stocke": fichier_stocke,
                    }
                    factures[num_facture] = rec
                else:
                    if ht_55 is not None:
                        rec["prix_HT_5_5pct"] = round((rec.get("prix_HT_5_5pct") or 0.0) + ht_55, 2)
                    if ht_10 is not None:
                        rec["prix_HT_10pct"] = round((rec.get("prix_HT_10pct") or 0.0) + ht_10, 2)
                    if ht_20 is not None:
                        rec["prix_HT_20pct"] = round((rec.get("prix_HT_20pct") or 0.0) + ht_20, 2)
                    if not rec.get("date_emission"):
                        rec["date_emission"] = date_emission
                    if not rec.get("date_paiement_prevue"):
                        rec["date_paiement_prevue"] = date_paiement
                    if commentaire and not rec.get("fichier_source"):
                        rec["fichier_source"] = commentaire
                    if commentaire and not rec.get("fichier_stocke"):
                        if os.path.exists(os.path.join(storage_dir, commentaire)):
                            rec["fichier_stocke"] = commentaire

            # BL
            if num_bl:
                if num_bl not in bons:
                    fichier_stocke = None
                    if commentaire and os.path.exists(os.path.join(storage_dir, commentaire)):
                        fichier_stocke = commentaire
                    bons[num_bl] = {
                        "numero_bon_livraison": num_bl,
                        "nom_fournisseur": fkey,
                        "date_livraison": date_emission,
                        "prix_HT_5_5pct": ht_55,
                        "prix_HT_10pct": ht_10,
                        "prix_HT_20pct": ht_20,
                        "numero_facture_rattachee": num_facture,
                        "fichier_source": commentaire or "",
                        "fichier_stocke": fichier_stocke,
                    }
                else:
                    bon = bons[num_bl]
                    if bon.get("numero_facture_rattachee") is None and num_facture:
                        bon["numero_facture_rattachee"] = num_facture
                    if not bon.get("date_livraison"):
                        bon["date_livraison"] = date_emission
                    if commentaire and not bon.get("fichier_source"):
                        bon["fichier_source"] = commentaire
                    if commentaire and not bon.get("fichier_stocke"):
                        if os.path.exists(os.path.join(storage_dir, commentaire)):
                            bon["fichier_stocke"] = commentaire

            if num_facture and num_bl:
                lst = fac_to_bls.setdefault(num_facture, [])
                if num_bl not in lst:
                    lst.append(num_bl)
    finally:
        wb.close()

    # Insertion : factures d'abord, puis BL (FK)
    for rec in factures.values():
        repo.upsert_facture(rec)
    for rec in bons.values():
        repo.upsert_bon(rec)

    return (len(factures), len(bons))


# ---------------------------------------------------------------------------
# Seed Autres achats
# ---------------------------------------------------------------------------

def seed_autres_achats(xlsm_path: str) -> int:
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)
    if "Autres achats" not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb["Autres achats"]
    inserted = 0
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row):
                continue
            fournisseur = _to_str_or_none(row[0])  # A
            if not fournisseur:
                continue
            data = {
                "fournisseur": fournisseur,
                "categorie": _to_str_or_none(row[1]),  # B
                "num_facture": _to_str_or_none(row[2]),  # C
                "num_bl": _to_str_or_none(row[3]),     # D
                "date": _to_iso_date(row[4]),           # E
                # F=mois, G=année (formules dérivées) → ignorées
                "ht_0": _to_float_or_none(row[7]),    # H
                "ht_2_1": _to_float_or_none(row[8]),  # I
                "ht_5_5": _to_float_or_none(row[9]),  # J
                "ht_10": _to_float_or_none(row[10]),  # K
                "ht_20": _to_float_or_none(row[11]),  # L
                # M..S = formules (TVA, TTC) ignorées
                "conditions": _to_str_or_none(row[20]),     # U
                "date_paiement": _to_iso_date(row[21]),     # V
                # W..Y = formules check
                "amortissable": _to_str_or_none(row[25]),   # Z
                "ref_denotage": _to_str_or_none(row[26]) if len(row) > 26 else None,  # AA
            }
            repo.insert_autre_achat(data)
            inserted += 1
    finally:
        wb.close()
    return inserted


# ---------------------------------------------------------------------------
# Seed DOMINO (depuis le JSON existant)
# ---------------------------------------------------------------------------

def seed_domino(json_path: str = DEFAULT_DOMINO_JSON) -> int:
    if not os.path.exists(json_path):
        return 0
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        print(f"[SEED] Lecture {json_path} impossible: {e}")
        return 0

    if not isinstance(payload, dict):
        return 0

    inserted = 0
    for _, item in payload.items():
        if not isinstance(item, dict):
            continue
        data = item.get("data") if isinstance(item.get("data"), dict) else None
        if not data:
            continue
        try:
            repo.upsert_domino_jour({
                **data,
                "imported_at": item.get("imported_at"),
            })
            inserted += 1
        except Exception as e:
            print(f"[SEED] DOMINO entrée ignorée ({data.get('date')}): {e}")
    return inserted


# ---------------------------------------------------------------------------
# Orchestrateur
# ---------------------------------------------------------------------------

def seed_if_empty(xlsm_path: str | None = None) -> dict:
    """
    Seed chaque table de la BDD **indépendamment** si elle est vide.

    Contrairement à l'ancienne logique (tout-ou-rien), cette version vérifie
    table par table : une table déjà peuplée est ignorée, une table vide est
    seedée même si les autres ne le sont pas.

    Retourne un résumé des opérations avec le détail par table.
    """
    db.get_conn()  # init schéma

    empty = db.tables_empty_status()

    # Si absolument tout est déjà peuplé, on sort rapidement
    if not any(empty.values()):
        return {"seeded": False, "reason": "toutes les tables sont déjà peuplées"}

    path = xlsm_path or _xlsm_path()
    summary: dict = {
        "seeded": True,
        "xlsm_path": path,
        "fournisseurs": 0,
        "factures": 0,
        "bons": 0,
        "autres_achats": 0,
        "domino_jours": 0,
        "skipped": [],
    }

    # --- Fournisseurs (prérequis des autres tables) ---
    if empty["fournisseurs"]:
        summary["fournisseurs"] = seed_fournisseurs(path)
    else:
        summary["skipped"].append("fournisseurs")

    if path:
        # --- Factures + BL ---
        if empty["factures"] or empty["bons_livraison"]:
            try:
                nb_f, nb_b = seed_achats_cons(path)
                summary["factures"] = nb_f
                summary["bons"] = nb_b
            except Exception as e:
                print(f"[SEED] Erreur Achats Cons: {e}")
        else:
            summary["skipped"].append("factures")
            summary["skipped"].append("bons_livraison")

        # --- Autres achats ---
        if empty["autres_achats"]:
            try:
                summary["autres_achats"] = seed_autres_achats(path)
            except Exception as e:
                print(f"[SEED] Erreur Autres achats: {e}")
        else:
            summary["skipped"].append("autres_achats")

    # --- DOMINO ---
    if empty["domino_jours"]:
        summary["domino_jours"] = seed_domino()
    else:
        summary["skipped"].append("domino_jours")

    return summary


if __name__ == "__main__":
    summary = seed_if_empty()
    print(json.dumps(summary, indent=2, ensure_ascii=False))
