from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Literal, Any, cast
from datetime import date, datetime, timedelta
from concurrent.futures import ThreadPoolExecutor
import asyncio
import tempfile
import shutil
import time
import json
import os
import re
import zipfile
import threading
import uuid

from main import (
    load_pdf_text,
    classify_document,
    build_prompt,
    finalize_document_data,
    link_documents,
    write_to_achats_cons,
    FOURNISSEUR_DISPLAY,
    llm,
)
import openpyxl
from openpyxl import load_workbook
import domino as domino_module

# Dossier de stockage persistant des PDFs importés
STORAGE_DIR = "storage"
os.makedirs(STORAGE_DIR, exist_ok=True)

app = FastAPI(
    title="Marjo — API Gestion Factures",
    description="API d'extraction automatique de factures et bons de livraison par IA",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Sert les PDFs stockés via /api/documents/<filename>
app.mount("/api/documents", StaticFiles(directory=STORAGE_DIR), name="documents")

# Pool de threads dédié au traitement LLM (bloquant)
_executor = ThreadPoolExecutor(max_workers=4)
_regen_lock = threading.Lock()
_xlsm_write_lock = threading.Lock()
_regen_pending = False
_regen_running = False
_domino_resync_jobs: dict[str, dict] = {}
_domino_resync_jobs_lock = threading.Lock()
_automation_tasks: dict[str, dict[str, Any]] = {}
_automation_logs: list[dict[str, Any]] = []
_automation_lock = threading.Lock()
_automation_scheduler_started = False

# ---------------------------------------------------------------------------
# Fichier source de vérité unique
# ---------------------------------------------------------------------------
TRESORERIE_XLSM = os.getenv(
    "TRESORERIE_XLSM_PATH",
    "output/Suivi trésorerie MLC.xlsm",
)
TRESORERIE_XLSM_FALLBACK = "output/Suivi trésorerie MLC - Copie.xlsm"


def _is_valid_xlsm(path: str) -> bool:
    if not os.path.exists(path):
        return False
    try:
        with zipfile.ZipFile(path, "r") as zf:
            names = zf.namelist()
            return "[Content_Types].xml" in names and zf.testzip() is None
    except (zipfile.BadZipFile, OSError):
        return False


def _resolve_tresorerie_path() -> str:
    if _is_valid_xlsm(TRESORERIE_XLSM):
        return TRESORERIE_XLSM
    if _is_valid_xlsm(TRESORERIE_XLSM_FALLBACK):
        print(
            f"[WARN] Fichier principal xlsm invalide: '{TRESORERIE_XLSM}'. "
            f"Utilisation du fallback '{TRESORERIE_XLSM_FALLBACK}'."
        )
        return TRESORERIE_XLSM_FALLBACK
    return TRESORERIE_XLSM


def _iter_tresorerie_candidates() -> list[str]:
    # Garder l'ordre principal -> fallback, sans doublons.
    out: list[str] = []
    for p in (TRESORERIE_XLSM, TRESORERIE_XLSM_FALLBACK):
        if p and p not in out:
            out.append(p)
    return out


def _pick_valid_tresorerie_path() -> str | None:
    for path in _iter_tresorerie_candidates():
        if _is_valid_xlsm(path):
            return path
    return None


def _build_tresorerie_invalid_detail() -> str:
    states: list[str] = []
    for path in _iter_tresorerie_candidates():
        exists = os.path.exists(path)
        valid = _is_valid_xlsm(path)
        size = os.path.getsize(path) if exists else 0
        states.append(f"{path} (exists={exists}, valid={valid}, size={size} bytes)")
    return "; ".join(states)


def _ensure_valid_tresorerie_path() -> str:
    path = _pick_valid_tresorerie_path()
    if path:
        return path

    any_exists = any(os.path.exists(p) for p in _iter_tresorerie_candidates())
    if any_exists:
        raise HTTPException(
            status_code=409,
            detail=(
                "Aucun fichier xlsm valide trouve pour l'export. "
                f"Etat: {_build_tresorerie_invalid_detail()}"
            ),
        )
    raise HTTPException(
        status_code=404,
        detail=(
            "Fichier 'Suivi trésorerie MLC.xlsm' introuvable dans output/. "
            f"Chemins testes: {_build_tresorerie_invalid_detail()}"
        ),
    )


def _backup_path_for(path: str) -> str:
    return f"{path}.lastgood.bak"


def _restore_tresorerie_from_backup() -> dict:
    """Restaure le fichier trésorerie depuis sa backup .lastgood.bak."""
    candidates = _iter_tresorerie_candidates()

    for target in candidates:
        backup = _backup_path_for(target)
        if os.path.exists(backup) and _is_valid_xlsm(backup):
            os.makedirs(os.path.dirname(target) or ".", exist_ok=True)
            shutil.copy2(backup, target)
            if not _is_valid_xlsm(target):
                raise HTTPException(
                    status_code=500,
                    detail=(
                        "La restauration de backup a echoue: le fichier restaure est invalide. "
                        f"target={target}, backup={backup}"
                    ),
                )
            return {
                "message": "Restauration XLSM effectuee depuis la backup last-good.",
                "target": target,
                "backup": backup,
            }

    raise HTTPException(
        status_code=404,
        detail=(
            "Aucune backup XLSM valide trouvee (.lastgood.bak). "
            "Faites au moins une ecriture reussie pour generer une backup."
        ),
    )


# ---------------------------------------------------------------------------
# Automatisation — tâches planifiées + logs
# ---------------------------------------------------------------------------

def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _add_automation_log(task_id: str, level: Literal["info", "warn", "error"], message: str, details: dict | None = None) -> None:
    entry = {
        "timestamp": _now_iso(),
        "task_id": task_id,
        "level": level,
        "message": message,
        "details": details or {},
    }
    with _automation_lock:
        _automation_logs.append(entry)
        # Garder l'historique borné en mémoire
        if len(_automation_logs) > 2000:
            del _automation_logs[: len(_automation_logs) - 2000]


def _init_automation_tasks() -> None:
    with _automation_lock:
        if _automation_tasks:
            return
        now = datetime.now()
        _automation_tasks["mail_fetch"] = {
            "id": "mail_fetch",
            "label": "Recuperation des mails",
            "description": "Simulation de récupération de pièces jointes mails.",
            "interval_seconds": 300,
            "enabled": False,
            "is_running": False,
            "last_start": None,
            "last_end": None,
            "last_status": "idle",
            "last_error": None,
            "run_count": 0,
            "error_count": 0,
            "next_run": now.isoformat(timespec="seconds"),
        }
        _automation_tasks["domino_auto_import"] = {
            "id": "domino_auto_import",
            "label": "Import DOMINO automatique",
            "description": "Importe les nouveaux fichiers DOMINO non traités dans XLSM.",
            "interval_seconds": 600,
            "enabled": False,
            "is_running": False,
            "last_start": None,
            "last_end": None,
            "last_status": "idle",
            "last_error": None,
            "run_count": 0,
            "error_count": 0,
            "next_run": now.isoformat(timespec="seconds"),
        }
        _automation_tasks["xlsx_healthcheck"] = {
            "id": "xlsx_healthcheck",
            "label": "Controle sante XLSM",
            "description": "Vérifie la validité du fichier XLSM principal/fallback.",
            "interval_seconds": 900,
            "enabled": False,
            "is_running": False,
            "last_start": None,
            "last_end": None,
            "last_status": "idle",
            "last_error": None,
            "run_count": 0,
            "error_count": 0,
            "next_run": now.isoformat(timespec="seconds"),
        }


def _run_mail_fetch_task() -> dict:
    # Placeholder robuste: on log l'état des dossiers d'entrée/sortie.
    storage_files = len(os.listdir(STORAGE_DIR)) if os.path.exists(STORAGE_DIR) else 0
    domino_files = len([f for f in os.listdir(domino_module.DOMINO_FOLDER)]) if os.path.exists(domino_module.DOMINO_FOLDER) else 0
    _add_automation_log(
        "mail_fetch",
        "info",
        "Vérification de récupération mails exécutée.",
        {"storage_files": storage_files, "domino_files": domino_files},
    )
    return {"storage_files": storage_files, "domino_files": domino_files}


def _run_domino_auto_import_task() -> dict:
    files = domino_module.list_domino_files()
    pending = [f for f in files if not f.get("imported")]
    if not pending:
        _add_automation_log("domino_auto_import", "info", "Aucun nouveau fichier DOMINO a importer.")
        return {"pending": 0, "imported": 0}

    imported_count = 0
    active_xlsm = _pick_valid_tresorerie_path()
    with _xlsm_write_lock:
        for f in pending:
            try:
                res = domino_module.import_domino_file(
                    filename=f["filename"],
                    xlsm_path=active_xlsm,
                    overwrite=False,
                )
                if not res.get("skipped"):
                    imported_count += 1
            except Exception as e:
                _add_automation_log(
                    "domino_auto_import",
                    "error",
                    f"Echec import {f['filename']}",
                    {"error": str(e)},
                )

    _add_automation_log(
        "domino_auto_import",
        "info",
        "Import automatique DOMINO exécuté.",
        {"pending": len(pending), "imported": imported_count},
    )
    return {"pending": len(pending), "imported": imported_count}


def _run_xlsx_healthcheck_task() -> dict:
    states = []
    for p in _iter_tresorerie_candidates():
        exists = os.path.exists(p)
        valid = _is_valid_xlsm(p)
        size = os.path.getsize(p) if exists else 0
        states.append({"path": p, "exists": exists, "valid": valid, "size": size})
    _add_automation_log("xlsx_healthcheck", "info", "Contrôle XLSM exécuté.", {"files": states})
    return {"files": states}


def _execute_automation_task(task_id: str, trigger: Literal["manual", "scheduled"] = "manual") -> None:
    with _automation_lock:
        task = _automation_tasks.get(task_id)
        if not task:
            return
        if task.get("is_running"):
            return
        task["is_running"] = True
        task["last_start"] = _now_iso()
        task["last_status"] = "running"
        task["last_error"] = None

    _add_automation_log(task_id, "info", f"Execution de tache ({trigger}).")

    try:
        if task_id == "mail_fetch":
            result = _run_mail_fetch_task()
        elif task_id == "domino_auto_import":
            result = _run_domino_auto_import_task()
        elif task_id == "xlsx_healthcheck":
            result = _run_xlsx_healthcheck_task()
        else:
            raise ValueError(f"Tache inconnue: {task_id}")

        with _automation_lock:
            task = _automation_tasks[task_id]
            task["run_count"] += 1
            task["last_status"] = "ok"
            task["last_end"] = _now_iso()
            next_run_dt = datetime.now() + timedelta(seconds=int(task["interval_seconds"]))
            task["next_run"] = next_run_dt.isoformat(timespec="seconds")
            task["is_running"] = False
        _add_automation_log(task_id, "info", "Execution terminee.", {"result": result})
    except Exception as e:
        with _automation_lock:
            task = _automation_tasks[task_id]
            task["error_count"] += 1
            task["last_status"] = "error"
            task["last_error"] = str(e)
            task["last_end"] = _now_iso()
            next_run_dt = datetime.now() + timedelta(seconds=int(task["interval_seconds"]))
            task["next_run"] = next_run_dt.isoformat(timespec="seconds")
            task["is_running"] = False
        _add_automation_log(task_id, "error", "Execution en echec.", {"error": str(e)})


def _automation_scheduler_loop() -> None:
    global _automation_scheduler_started
    while True:
        due_ids: list[str] = []
        now = datetime.now()
        with _automation_lock:
            for task_id, task in _automation_tasks.items():
                if not task.get("enabled") or task.get("is_running"):
                    continue
                next_run = task.get("next_run")
                try:
                    next_dt = datetime.fromisoformat(next_run) if next_run else now
                except ValueError:
                    next_dt = now
                if now >= next_dt:
                    due_ids.append(task_id)

        for task_id in due_ids:
            _executor.submit(_execute_automation_task, task_id, "scheduled")

        # heartbeat léger
        time.sleep(1.0)


def _start_automation_scheduler_once() -> None:
    global _automation_scheduler_started
    with _automation_lock:
        if _automation_scheduler_started:
            return
        _automation_scheduler_started = True
    t = threading.Thread(target=_automation_scheduler_loop, daemon=True, name="automation-scheduler")
    t.start()

# ---------------------------------------------------------------------------
# Store en mémoire (remplaçable par une BDD)
# Clé d'unicité : numero_facture pour les factures, numero_bon_livraison pour les BL
# ---------------------------------------------------------------------------
_store: dict[str, dict[str, dict]] = {"factures": {}, "bons": {}}


def _recompute_derived(record: dict) -> dict:
    """
    Recalcule les champs dérivés (TVA, TTC, vérifications) depuis les bases HT.
    Modifie le dict en place et le retourne.
    """
    ht_55 = record.get("prix_HT_5_5pct")
    ht_10 = record.get("prix_HT_10pct")
    ht_20 = record.get("prix_HT_20pct")

    # Total HT
    ht_vals = [v for v in (ht_55, ht_10, ht_20) if v is not None]
    record["montant_total"] = round(sum(ht_vals), 2) if ht_vals else None

    # TVA calculée
    tva_55 = round(ht_55 * 0.055, 2) if ht_55 is not None else None
    tva_10 = round(ht_10 * 0.1,   2) if ht_10 is not None else None
    tva_20 = round(ht_20 * 0.2,   2) if ht_20 is not None else None
    record["tva_5_5pct"] = tva_55
    record["tva_10pct"]  = tva_10
    record["tva_20pct"]  = tva_20

    # Total TVA
    tva_vals = [v for v in (tva_55, tva_10, tva_20) if v is not None]
    total_tva = round(sum(tva_vals), 2) if tva_vals else None
    record["total_tva"] = total_tva

    # TTC
    tot_ht = record["montant_total"]
    if tot_ht is not None or total_tva is not None:
        record["montant_ttc"] = round((tot_ht or 0) + (total_tva or 0), 2)
    else:
        record["montant_ttc"] = None

    # Vérifications
    def _vf(ht, tva, rate):
        if ht is None or tva is None or ht == 0:
            return ""
        return "OK" if round(tva / ht, 3) == rate else "Erreur"

    record["verif_tva_5_5"] = _vf(ht_55, tva_55, 0.055)
    record["verif_tva_10"]  = _vf(ht_10, tva_10, 0.1)
    record["verif_tva_20"]  = _vf(ht_20, tva_20, 0.2)

    return record


@app.on_event("startup")
def _startup_load_excel() -> None:
    """
    Au démarrage, recharge le store depuis l'onglet 'Achats Cons' du fichier
    'Suivi trésorerie MLC.xlsm' en mode ligne-à-ligne.

    Cas gérés :
    - ligne facture sans BL : D rempli, E vide
    - ligne BL rattaché à facture : D rempli, E rempli
    - ligne BL sans facture : D vide, E rempli

    Les factures multi-BL sont agrégées par numero_facture pour les montants HT.
    """
    active_xlsm = _pick_valid_tresorerie_path()
    if not active_xlsm:
        print(
            "[WARN] Aucun xlsm valide au demarrage - store vide. "
            f"Etat: {_build_tresorerie_invalid_detail()}"
        )
        return

    try:
        # data_only=True pour lire les valeurs calculées des formules Excel (cache)
        wb = load_workbook(active_xlsm, read_only=True, data_only=True)
        ws = wb["Achats Cons"]

        # Mapping fournisseur affiché -> clé interne
        display_to_key = _get_display_to_key()

        for row in ws.iter_rows(min_row=2, values_only=True):
            fournisseur_raw = row[2]   # col C
            if not fournisseur_raw:
                continue
            fournisseur_display = str(fournisseur_raw).strip()
            fournisseur_key = display_to_key.get(fournisseur_display.lower())
            if not fournisseur_key:
                # Aucun filtre fournisseur : on enregistre automatiquement
                # les fournisseurs inconnus trouvés dans le fichier.
                proposed = _make_supplier_key(fournisseur_display)
                fournisseur_key = proposed
                idx = 2
                while (
                    fournisseur_key in _fournisseurs
                    and _fournisseurs[fournisseur_key]["nom_affiche"].lower() != fournisseur_display.lower()
                ):
                    fournisseur_key = f"{proposed}_{idx}"
                    idx += 1

                if fournisseur_key not in _fournisseurs:
                    _fournisseurs[fournisseur_key] = {
                        "id": fournisseur_key,
                        "nom_affiche": fournisseur_display,
                        "patterns": [fournisseur_display.lower()],
                    }

                display_to_key[fournisseur_display.lower()] = fournisseur_key

            num_facture  = str(row[3]).strip() if row[3] is not None else None
            num_bl       = str(row[4]).strip() if row[4] is not None else None
            date_emission = row[5]   # col F — datetime ou None
            ht_55        = row[8]    # col I
            ht_10        = row[9]    # col J
            ht_20        = row[10]   # col K
            date_paiement = row[18]  # col S
            commentaire  = str(row[22]).strip() if row[22] is not None else None  # col W

            # Convertir les datetime Excel en date Python
            def _to_date(v):
                if v is None:
                    return None
                if hasattr(v, "date"):
                    return v.date()
                if isinstance(v, date):
                    return v
                return None

            def _to_float(v):
                if v is None:
                    return None
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return None

            # Une ligne utile doit contenir une facture ou un BL.
            if not num_facture and not num_bl:
                continue

            ht_55_v = _to_float(ht_55)
            ht_10_v = _to_float(ht_10)
            ht_20_v = _to_float(ht_20)

            # Construire ou enrichir l'entrée facture (si la ligne porte une facture)
            if num_facture and num_facture not in _store["factures"]:
                record: dict = {
                    "type_document":       "facture",
                    "numero_facture":      num_facture,
                    "nom_fournisseur":     fournisseur_key,
                    "date_emission":       _to_date(date_emission),
                    "date_paiement_prevue": _to_date(date_paiement),
                    "prix_HT_5_5pct":      ht_55_v,
                    "prix_HT_10pct":       ht_10_v,
                    "prix_HT_20pct":       ht_20_v,
                    "montant_total":       None,
                    "tva_5_5pct":          None,
                    "tva_10pct":           None,
                    "tva_20pct":           None,
                    "total_tva":           None,
                    "montant_ttc":         None,
                    "verif_tva_5_5":       "",
                    "verif_tva_10":        "",
                    "verif_tva_20":        "",
                    "bons_livraisons":     [],
                    "fichier_source":      commentaire or "",
                    "fichier_stocke":      None,
                }
                # Vérifier si le PDF est en storage
                if commentaire:
                    candidate = os.path.join(STORAGE_DIR, commentaire)
                    if os.path.exists(candidate):
                        record["fichier_stocke"] = commentaire
                _recompute_derived(record)
                _store["factures"][num_facture] = record
            elif num_facture:
                # Ligne supplémentaire pour la même facture (plusieurs BL) :
                # on agrège les HT par taux.
                record = _store["factures"][num_facture]
                if ht_55_v is not None:
                    record["prix_HT_5_5pct"] = round((record.get("prix_HT_5_5pct") or 0.0) + ht_55_v, 2)
                if ht_10_v is not None:
                    record["prix_HT_10pct"] = round((record.get("prix_HT_10pct") or 0.0) + ht_10_v, 2)
                if ht_20_v is not None:
                    record["prix_HT_20pct"] = round((record.get("prix_HT_20pct") or 0.0) + ht_20_v, 2)
                if not record.get("date_emission"):
                    record["date_emission"] = _to_date(date_emission)
                if not record.get("date_paiement_prevue"):
                    record["date_paiement_prevue"] = _to_date(date_paiement)
                if commentaire and not record.get("fichier_source"):
                    record["fichier_source"] = commentaire
                if commentaire and not record.get("fichier_stocke"):
                    candidate = os.path.join(STORAGE_DIR, commentaire)
                    if os.path.exists(candidate):
                        record["fichier_stocke"] = commentaire
                _recompute_derived(record)

            # Rattacher le BL à la facture
            if num_facture and num_bl and num_bl not in _store["factures"][num_facture]["bons_livraisons"]:
                _store["factures"][num_facture]["bons_livraisons"].append(num_bl)

            # Créer/enrichir l'entrée BL si présent
            if num_bl and num_bl not in _store["bons"]:
                _store["bons"][num_bl] = {
                    "type_document":           "bon_livraison",
                    "numero_bon_livraison":    num_bl,
                    "nom_fournisseur":         fournisseur_key,
                    "date_livraison":          _to_date(date_emission),  # col F = date du BL
                    "prix_HT_5_5pct":          ht_55_v,
                    "prix_HT_10pct":           ht_10_v,
                    "prix_HT_20pct":           ht_20_v,
                    "montant_total":           None,
                    "tva_5_5pct":              None,
                    "tva_10pct":               None,
                    "tva_20pct":               None,
                    "total_tva":               None,
                    "montant_ttc":             None,
                    "verif_tva_5_5":           "",
                    "verif_tva_10":            "",
                    "verif_tva_20":            "",
                    "numero_facture_rattachee": num_facture,
                    "fichier_source":          commentaire or "",
                    "fichier_stocke":          None,
                }
                if commentaire:
                    candidate = os.path.join(STORAGE_DIR, commentaire)
                    if os.path.exists(candidate):
                        _store["bons"][num_bl]["fichier_stocke"] = commentaire
                _recompute_derived(_store["bons"][num_bl])
            elif num_bl:
                bon = _store["bons"][num_bl]
                if bon.get("numero_facture_rattachee") is None and num_facture:
                    bon["numero_facture_rattachee"] = num_facture
                if not bon.get("date_livraison"):
                    bon["date_livraison"] = _to_date(date_emission)
                if commentaire and not bon.get("fichier_source"):
                    bon["fichier_source"] = commentaire
                if commentaire and not bon.get("fichier_stocke"):
                    candidate = os.path.join(STORAGE_DIR, commentaire)
                    if os.path.exists(candidate):
                        bon["fichier_stocke"] = commentaire
                _recompute_derived(bon)

        # Sérialiser les dates en strings ISO pour le store
        for f in _store["factures"].values():
            for field in ("date_emission", "date_paiement_prevue"):
                v = f.get(field)
                if isinstance(v, date) and not isinstance(v, str):
                    f[field] = v.isoformat()
        for b in _store["bons"].values():
            v = b.get("date_livraison")
            if isinstance(v, date) and not isinstance(v, str):
                b["date_livraison"] = v.isoformat()

        nb_f = len(_store["factures"])
        nb_b = len(_store["bons"])
        print(f"[OK] Store rechargé depuis '{active_xlsm}' : {nb_f} facture(s), {nb_b} bon(s) de livraison.")

    except Exception as e:
        print(f"[WARN] Impossible de charger le fichier xlsm au démarrage : {e}")


@app.on_event("startup")
def _startup_domino_auto_import() -> None:
    """
    Au démarrage, importe automatiquement les fichiers DOMINO non encore traités.
    N'écrase jamais les imports existants.
    """
    files = domino_module.list_domino_files()
    pending = [f for f in files if not f["imported"]]
    if not pending:
        return

    active_xlsm = _pick_valid_tresorerie_path()
    imported_count = 0
    for f in pending:
        try:
            result = domino_module.import_domino_file(
                filename=f["filename"],
                xlsm_path=active_xlsm,
                overwrite=False,
            )
            if not result.get("skipped"):
                imported_count += 1
                print(f"[DOMINO] Auto-import '{f['filename']}' : {result['message']}")
        except Exception as e:
            print(f"[DOMINO] Erreur auto-import '{f['filename']}' : {e}")

    if imported_count:
        print(f"[DOMINO] {imported_count} fichier(s) auto-importé(s) au démarrage.")


@app.on_event("startup")
def _startup_automation_scheduler() -> None:
    _init_automation_tasks()
    _start_automation_scheduler_once()
    print("[AUTOMATION] Scheduler initialisé.")


def _serialize(obj):
    """Convertit les objets date en string ISO pour la sérialisation JSON."""
    if isinstance(obj, date):
        return obj.isoformat()
    return obj


def _serialize_record(record: dict) -> dict:
    return {k: _serialize(v) for k, v in record.items()}


def _upsert_facture(data: dict) -> tuple[dict, str]:
    """
    Insère ou remplace une facture dans le store.
    Retourne (data, action) où action = 'created' | 'updated' | 'rejected'.
    Bloque si numero_facture est null.
    """
    numero = data.get("numero_facture")
    if not numero:
        return data, "rejected"
    action = "updated" if numero in _store["factures"] else "created"
    _recompute_derived(data)
    _store["factures"][numero] = data
    return data, action


def _upsert_bon(data: dict) -> tuple[dict, str]:
    """
    Insère ou remplace un bon de livraison dans le store.
    Retourne (data, action) où action = 'created' | 'updated' | 'rejected'.
    Bloque si numero_bon_livraison est null.
    """
    numero = data.get("numero_bon_livraison")
    if not numero:
        return data, "rejected"
    action = "updated" if numero in _store["bons"] else "created"
    _recompute_derived(data)
    _store["bons"][numero] = data
    return data, action


# ---------------------------------------------------------------------------
# Store fournisseurs — initialisé avec les 3 fournisseurs historiques
# Clé = identifiant interne (ex: "SYSCO"), valeur = dict avec nom_affiche + patterns
# ---------------------------------------------------------------------------
_fournisseurs: dict[str, dict] = {
    "SYSCO": {
        "id":           "SYSCO",
        "nom_affiche":  "Sysco",
        "patterns":     ["sysco"],
    },
    "AMBELYS": {
        "id":           "AMBELYS",
        "nom_affiche":  "Ambelys",
        "patterns":     ["ambelys"],
    },
    "TERREAZUR": {
        "id":           "TERREAZUR",
        "nom_affiche":  "TerreAzur",
        "patterns":     ["terreazur", "terre azur"],
    },
}


def _get_fournisseur_display() -> dict[str, str]:
    """Retourne {id → nom_affiche} pour tous les fournisseurs."""
    return {k: v["nom_affiche"] for k, v in _fournisseurs.items()}


def _get_display_to_key() -> dict[str, str]:
    """Retourne {nom_affiche.lower() → id} pour la lecture du xlsm."""
    return {v["nom_affiche"].lower(): k for k, v in _fournisseurs.items()}


def _make_supplier_key(display_name: str) -> str:
    """Construit une clé fournisseur stable depuis le libellé Excel."""
    base = re.sub(r"[^A-Z0-9]+", "_", display_name.upper()).strip("_")
    return base or "FOURNISSEUR_INCONNU"


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.get("/health")
def health():
    return {"status": "ok"}


def _process_one_pdf(tmp_path: str, filename: str, fournisseur_ids: list[str]) -> dict:
    """
    Traitement complet d'un PDF (bloquant — exécuté dans le thread pool).
    Retourne un dict avec les clés : data, doc_type, error.
    fournisseur_ids : liste des identifiants internes connus (ex: ["SYSCO", "AMBELYS", ...])
    """
    try:
        text     = load_pdf_text(tmp_path)
        doc_type = classify_document(text, filename)
        prompt   = build_prompt(doc_type, text, fournisseur_ids=fournisseur_ids)

        result = llm.invoke(prompt)
        model_dump = getattr(result, "model_dump", None)
        if callable(model_dump):
            data = cast(dict[str, Any], model_dump())
        elif isinstance(result, dict):
            data = cast(dict[str, Any], result)
        else:
            data = cast(dict[str, Any], dict(result))
        data   = finalize_document_data(
            data, text=text, filename=filename, predicted_type=doc_type,
            fournisseur_patterns={k: v["patterns"] for k, v in _fournisseurs.items()},
        )
        data   = _serialize_record(data)
        return {"data": data, "doc_type": doc_type, "error": None}
    except Exception as e:
        return {"data": None, "doc_type": None, "error": str(e)}


@app.post("/api/upload", summary="Uploader et analyser des PDFs")
async def upload_documents(files: list[UploadFile] = File(...)):
    """
    Reçoit un ou plusieurs fichiers PDF, les analyse via l'IA et retourne
    les données extraites.

    - Traitement parallèle dans un ThreadPoolExecutor (évite le socket hang up).
    - Unicité : numero_facture / numero_bon_livraison.
    - Numéro null → rejeté. Numéro existant → mis à jour (upsert).
    """
    if llm is None:
        raise HTTPException(
            status_code=503,
            detail="Service IA non configure (variables APIM_OPENAI_* manquantes).",
        )

    if not files:
        raise HTTPException(status_code=400, detail="Aucun fichier fourni.")

    results = {
        "created":  {"factures": 0, "bons": 0},
        "updated":  {"factures": 0, "bons": 0},
        "rejected": [],
        "errors":   [],
        "records":  [],  # liste des records extraits avec leur type et action
    }
    tmp_dir = tempfile.mkdtemp()
    loop    = asyncio.get_event_loop()

    try:
        # 1. Sauvegarde des fichiers : tmp_dir pour le traitement LLM,
        #    storage/ pour la persistance définitive
        saved: list[tuple[str, str, str]] = []  # (tmp_path, storage_path, filename)
        for upload in files:
            fname = upload.filename or ""
            if not fname.lower().endswith(".pdf"):
                results["errors"].append({
                    "fichier": fname,
                    "erreur": "Seuls les fichiers PDF sont acceptés.",
                })
                continue

            content = await upload.read()

            # Chemin temporaire pour le traitement LLM
            tmp_path = os.path.join(tmp_dir, fname)
            with open(tmp_path, "wb") as f:
                f.write(content)

            # Chemin de stockage définitif (on écrase si même nom = même document)
            storage_path = os.path.join(STORAGE_DIR, fname)
            with open(storage_path, "wb") as f:
                f.write(content)

            saved.append((tmp_path, storage_path, fname))

        # 2. Traitement LLM en parallèle dans le thread pool
        fournisseur_ids = list(_fournisseurs.keys())
        futures = [
            loop.run_in_executor(_executor, _process_one_pdf, tmp_path, fname, fournisseur_ids)
            for tmp_path, _, fname in saved
        ]
        outcomes = await asyncio.gather(*futures)

        # 3. Intégration des résultats dans le store
        for (_, storage_path, fname), outcome in zip(saved, outcomes):
            if outcome["error"]:
                results["errors"].append({"fichier": fname, "erreur": outcome["error"]})
                continue

            data     = outcome["data"]
            doc_type = outcome["doc_type"]

            # Stocke le nom du fichier PDF pour pouvoir le servir
            data["fichier_stocke"] = fname

            if doc_type == "bon_livraison":
                record, action = _upsert_bon(_deserialize_record(data))
            else:
                record, action = _upsert_facture(_deserialize_record(data))

            if action == "rejected":
                results["rejected"].append({
                    "fichier": fname,
                    "type":    doc_type,
                    "raison":  "Numéro non extrait par l'IA (null).",
                })
            elif doc_type == "bon_livraison":
                results[action]["bons"] += 1
                results["records"].append({
                    "type":   "bon_livraison",
                    "action": action,
                    "data":   _serialize_record(record),
                })
            else:
                results[action]["factures"] += 1
                results["records"].append({
                    "type":   "facture",
                    "action": action,
                    "data":   _serialize_record(record),
                })

        # 4. Reliaison automatique BL ↔ Factures
        _relink_store()

        # 5. Régénération Excel
        _schedule_regenerate_excel()

        nb_traites = (
            results["created"]["factures"] + results["created"]["bons"]
            + results["updated"]["factures"] + results["updated"]["bons"]
        )
        return {
            "traites":  nb_traites,
            "created":  results["created"],
            "updated":  results["updated"],
            "rejetes":  results["rejected"],
            "erreurs":  results["errors"],
            "records":  results["records"],
            "factures": results["created"]["factures"] + results["updated"]["factures"],
            "bons":     results["created"]["bons"]     + results["updated"]["bons"],
        }

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@app.get("/api/factures", summary="Lister les factures extraites")
def get_factures():
    return [_serialize_record(f) for f in _store["factures"].values()]


@app.get("/api/bons-livraison", summary="Lister les bons de livraison extraits")
def get_bons_livraison():
    return [_serialize_record(b) for b in _store["bons"].values()]


class RattachementBL(BaseModel):
    numero_bon_livraison: str

class RattachementFacture(BaseModel):
    numero_facture: str

# Champs éditables par type de document
FACTURE_EDITABLE_FIELDS = {
    "date_emission", "date_paiement_prevue",
    "prix_HT_5_5pct", "prix_HT_10pct", "prix_HT_20pct",
    "numero_facture", "nom_fournisseur",
}
BON_EDITABLE_FIELDS = {
    "date_livraison",
    "prix_HT_5_5pct", "prix_HT_10pct", "prix_HT_20pct",
    "numero_bon_livraison", "nom_fournisseur",
}

class PatchFacture(BaseModel):
    date_emission:        str | None = None
    date_paiement_prevue: str | None = None
    montant_total:        float | None = None
    prix_HT_5_5pct:       float | None = None
    prix_HT_10pct:        float | None = None
    prix_HT_20pct:        float | None = None
    numero_facture:       str | None = None
    nom_fournisseur:      str | None = None
    conditions_paiement:  str | None = None

class PatchBon(BaseModel):
    date_livraison:       str | None = None
    montant_total:        float | None = None
    prix_HT_5_5pct:       float | None = None
    prix_HT_10pct:        float | None = None
    prix_HT_20pct:        float | None = None
    numero_bon_livraison: str | None = None
    nom_fournisseur:      str | None = None


@app.patch("/api/factures/{numero_facture}", summary="Modifier les champs d'une facture")
def patch_facture(numero_facture: str, body: PatchFacture):
    """
    Met à jour les champs éditables d'une facture.
    Seuls les champs non-None dans le body sont modifiés.
    Si numero_facture change, la clé du store est mise à jour.
    """
    facture = _store["factures"].get(numero_facture)
    if not facture:
        raise HTTPException(status_code=404, detail=f"Facture '{numero_facture}' introuvable.")

    updates = {k: v for k, v in body.model_dump().items() if v is not None}

    # Validation des dates
    for date_field in ("date_emission", "date_paiement_prevue"):
        if date_field in updates:
            try:
                date.fromisoformat(updates[date_field])
            except ValueError:
                raise HTTPException(status_code=422, detail=f"Format de date invalide pour '{date_field}' (attendu : YYYY-MM-DD).")

    # Validation fournisseur
    if "nom_fournisseur" in updates and updates["nom_fournisseur"] not in _fournisseurs:
        raise HTTPException(status_code=422, detail="nom_fournisseur inconnu dans la liste des fournisseurs configurés.")

    nouveau_numero = updates.pop("numero_facture", None)
    facture.update(updates)
    _recompute_derived(facture)

    # Changement de numéro : réindexation
    if nouveau_numero and nouveau_numero != numero_facture:
        if nouveau_numero in _store["factures"]:
            raise HTTPException(status_code=409, detail=f"La facture '{nouveau_numero}' existe déjà.")
        facture["numero_facture"] = nouveau_numero
        del _store["factures"][numero_facture]
        _store["factures"][nouveau_numero] = facture
    else:
        _store["factures"][numero_facture] = facture

    _schedule_regenerate_excel()
    return _serialize_record(facture)


@app.patch("/api/bons-livraison/{numero_bl}", summary="Modifier les champs d'un bon de livraison")
def patch_bon(numero_bl: str, body: PatchBon):
    """
    Met à jour les champs éditables d'un bon de livraison.
    Seuls les champs non-None dans le body sont modifiés.
    """
    bon = _store["bons"].get(numero_bl)
    if not bon:
        raise HTTPException(status_code=404, detail=f"Bon de livraison '{numero_bl}' introuvable.")

    updates = {k: v for k, v in body.model_dump().items() if v is not None}

    # Validation date
    if "date_livraison" in updates:
        try:
            date.fromisoformat(updates["date_livraison"])
        except ValueError:
            raise HTTPException(status_code=422, detail="Format de date invalide pour 'date_livraison' (attendu : YYYY-MM-DD).")

    # Validation fournisseur
    if "nom_fournisseur" in updates and updates["nom_fournisseur"] not in _fournisseurs:
        raise HTTPException(status_code=422, detail="nom_fournisseur inconnu dans la liste des fournisseurs configurés.")

    nouveau_numero = updates.pop("numero_bon_livraison", None)
    bon.update(updates)
    _recompute_derived(bon)

    # Changement de numéro : réindexation
    if nouveau_numero and nouveau_numero != numero_bl:
        if nouveau_numero in _store["bons"]:
            raise HTTPException(status_code=409, detail=f"Le bon '{nouveau_numero}' existe déjà.")
        bon["numero_bon_livraison"] = nouveau_numero
        del _store["bons"][numero_bl]
        _store["bons"][nouveau_numero] = bon
    else:
        _store["bons"][numero_bl] = bon

    _schedule_regenerate_excel()
    return _serialize_record(bon)


@app.patch("/api/factures/{numero_facture}/rattacher", summary="Rattacher un BL à une facture")
def rattacher_bl_a_facture(numero_facture: str, body: RattachementBL):
    """
    Rattache manuellement un BL à une facture :
    - Ajoute le BL dans bons_livraisons de la facture (si absent)
    - Met à jour numero_facture_rattachee sur le BL
    """
    facture = _store["factures"].get(numero_facture)
    if not facture:
        raise HTTPException(status_code=404, detail=f"Facture '{numero_facture}' introuvable.")

    bon = _store["bons"].get(body.numero_bon_livraison)
    if not bon:
        raise HTTPException(status_code=404, detail=f"Bon de livraison '{body.numero_bon_livraison}' introuvable.")

    # Mise à jour de la facture
    bons_list: list = facture.get("bons_livraisons") or []
    if body.numero_bon_livraison not in bons_list:
        bons_list.append(body.numero_bon_livraison)
    facture["bons_livraisons"] = bons_list
    _store["factures"][numero_facture] = facture

    # Mise à jour du BL
    bon["numero_facture_rattachee"] = numero_facture
    _store["bons"][body.numero_bon_livraison] = bon

    _schedule_regenerate_excel()
    return {
        "facture": _serialize_record(facture),
        "bon":     _serialize_record(bon),
    }


@app.patch("/api/bons-livraison/{numero_bl}/rattacher", summary="Rattacher une facture à un BL")
def rattacher_facture_a_bl(numero_bl: str, body: RattachementFacture):
    """
    Rattache manuellement une facture à un BL (symétrique de l'endpoint précédent).
    """
    bon = _store["bons"].get(numero_bl)
    if not bon:
        raise HTTPException(status_code=404, detail=f"Bon de livraison '{numero_bl}' introuvable.")

    facture = _store["factures"].get(body.numero_facture)
    if not facture:
        raise HTTPException(status_code=404, detail=f"Facture '{body.numero_facture}' introuvable.")

    # Mise à jour du BL
    bon["numero_facture_rattachee"] = body.numero_facture
    _store["bons"][numero_bl] = bon

    # Mise à jour de la facture
    bons_list: list = facture.get("bons_livraisons") or []
    if numero_bl not in bons_list:
        bons_list.append(numero_bl)
    facture["bons_livraisons"] = bons_list
    _store["factures"][body.numero_facture] = facture

    _schedule_regenerate_excel()
    return {
        "bon":     _serialize_record(bon),
        "facture": _serialize_record(facture),
    }


@app.delete("/api/factures/{numero_facture}/rattacher/{numero_bl}", summary="Supprimer le rattachement BL ↔ Facture")
def supprimer_rattachement(numero_facture: str, numero_bl: str):
    """Supprime le lien entre une facture et un BL (dans les deux sens)."""
    facture = _store["factures"].get(numero_facture)
    if not facture:
        raise HTTPException(status_code=404, detail=f"Facture '{numero_facture}' introuvable.")

    bon = _store["bons"].get(numero_bl)

    # Retrait du BL de la liste de la facture
    bons_list = facture.get("bons_livraisons") or []
    facture["bons_livraisons"] = [b for b in bons_list if b != numero_bl]
    _store["factures"][numero_facture] = facture

    # Retrait de la référence facture sur le BL
    if bon and bon.get("numero_facture_rattachee") == numero_facture:
        bon["numero_facture_rattachee"] = None
        _store["bons"][numero_bl] = bon

    _schedule_regenerate_excel()
    return {"message": f"Rattachement {numero_bl} ↔ {numero_facture} supprimé."}


# ---------------------------------------------------------------------------
# Endpoints fournisseurs
# ---------------------------------------------------------------------------

class FournisseurCreate(BaseModel):
    id:          str   # identifiant interne unique, ex: "METRO"
    nom_affiche: str   # nom affiché dans le xlsm, ex: "Metro"
    patterns:    list[str] = []  # mots-clés pour la détection auto dans les PDFs

class FournisseurUpdate(BaseModel):
    nom_affiche: str | None = None
    patterns:    list[str] | None = None


@app.get("/api/fournisseurs", summary="Lister les fournisseurs")
def get_fournisseurs():
    return list(_fournisseurs.values())


@app.post("/api/fournisseurs", summary="Ajouter un fournisseur", status_code=201)
def create_fournisseur(body: FournisseurCreate):
    key = body.id.upper().strip().replace(" ", "")
    if key in _fournisseurs:
        raise HTTPException(status_code=409, detail=f"Le fournisseur '{key}' existe déjà.")
    if not body.nom_affiche.strip():
        raise HTTPException(status_code=422, detail="nom_affiche ne peut pas être vide.")
    _fournisseurs[key] = {
        "id":          key,
        "nom_affiche": body.nom_affiche.strip(),
        "patterns":    [p.lower().strip() for p in body.patterns if p.strip()],
    }
    return _fournisseurs[key]


@app.patch("/api/fournisseurs/{fournisseur_id}", summary="Modifier un fournisseur")
def update_fournisseur(fournisseur_id: str, body: FournisseurUpdate):
    key = fournisseur_id.upper().strip()
    if key not in _fournisseurs:
        raise HTTPException(status_code=404, detail=f"Fournisseur '{key}' introuvable.")
    if body.nom_affiche is not None:
        if not body.nom_affiche.strip():
            raise HTTPException(status_code=422, detail="nom_affiche ne peut pas être vide.")
        _fournisseurs[key]["nom_affiche"] = body.nom_affiche.strip()
    if body.patterns is not None:
        _fournisseurs[key]["patterns"] = [p.lower().strip() for p in body.patterns if p.strip()]
    return _fournisseurs[key]


@app.delete("/api/fournisseurs/{fournisseur_id}", summary="Supprimer un fournisseur")
def delete_fournisseur(fournisseur_id: str):
    key = fournisseur_id.upper().strip()
    if key not in _fournisseurs:
        raise HTTPException(status_code=404, detail=f"Fournisseur '{key}' introuvable.")
    # Vérifier qu'aucune facture ne référence ce fournisseur
    en_cours = [
        f["numero_facture"] for f in _store["factures"].values()
        if f.get("nom_fournisseur") == key
    ]
    if en_cours:
        raise HTTPException(
            status_code=409,
            detail=f"Impossible de supprimer '{key}' : {len(en_cours)} facture(s) lui sont rattachées.",
        )
    del _fournisseurs[key]
    return {"message": f"Fournisseur '{key}' supprimé."}


@app.get("/api/stats", summary="Statistiques globales")
def get_stats():
    factures = list(_store["factures"].values())
    bons = list(_store["bons"].values())

    # Montant total = somme des HT de toutes les factures
    # (on additionne les 3 taux HT disponibles sur chaque facture)
    montant_total = 0.0
    for f in factures:
        for field in ("prix_HT_5_5pct", "prix_HT_10pct", "prix_HT_20pct"):
            v = f.get(field)
            if v is not None:
                try:
                    montant_total += float(v)
                except (TypeError, ValueError):
                    pass
    # Ajouter aussi les montants portés par les BL (cas multi-BL où la facture
    # n'a pas de montant propre mais chaque BL en a un)
    factures_avec_ht = {
        num for num, f in _store["factures"].items()
        if any(f.get(k) for k in ("prix_HT_5_5pct", "prix_HT_10pct", "prix_HT_20pct"))
    }
    for b in bons:
        fac = b.get("numero_facture_rattachee")
        if not fac or fac not in factures_avec_ht:
            for field in ("prix_HT_5_5pct", "prix_HT_10pct", "prix_HT_20pct"):
                v = b.get(field)
                if v is not None:
                    try:
                        montant_total += float(v)
                    except (TypeError, ValueError):
                        pass

    # BL non rattachés = BL dont la facture n'existe pas dans le store
    # (uploadés sans rattachement automatique réussi)
    bl_non_rattaches = sum(
        1 for b in bons
        if not b.get("numero_facture_rattachee")
        or b.get("numero_facture_rattachee") not in _store["factures"]
    )

    return {
        "nb_factures": len(factures),
        "nb_bons": len(bons),
        "montant_total": round(montant_total, 2),
        "bl_non_rattaches": bl_non_rattaches,
    }


@app.get("/api/export/tresorerie/download", summary="Télécharger le fichier Suivi Trésorerie MLC.xlsm")
def download_tresorerie():
    """Télécharge le fichier Suivi trésorerie MLC.xlsm (source de vérité)."""
    path = _ensure_valid_tresorerie_path()
    return FileResponse(
        path=path,
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        filename="Suivi trésorerie MLC.xlsm",
    )


@app.post("/api/export/tresorerie", summary="Générer le fichier Suivi Trésorerie MLC.xlsm")
def export_tresorerie():
    """Force la régénération du xlsm depuis le store en mémoire."""
    active_xlsm = _ensure_valid_tresorerie_path()

    try:
        factures = list(_store["factures"].values())
        bons = list(_store["bons"].values())
        with _xlsm_write_lock:
            lignes_inserees = write_to_achats_cons(
                factures=factures,
                bons=bons,
                template_path=active_xlsm,
                output_path=active_xlsm,
                fournisseur_display=_get_fournisseur_display(),
            )
    except HTTPException:
        raise
    except Exception as e:
        err = f"{type(e).__name__}: {e}".rstrip(": ")
        raise HTTPException(status_code=500, detail=f"Export trésorerie impossible: {err}")

    return {
        "lignes_inserees": lignes_inserees,
        "fichier": os.path.basename(active_xlsm),
        "message": "Export trésorerie généré.",
    }


@app.post("/api/export/tresorerie/restore-lastgood", summary="Restaurer le XLSM depuis la backup last-good")
def restore_tresorerie_lastgood():
    """Restaure le fichier xlsm principal/fallback à partir de la backup .lastgood.bak."""
    with _xlsm_write_lock:
        return _restore_tresorerie_from_backup()


@app.delete("/api/reset", summary="Réinitialiser le store en mémoire (ne touche pas au xlsm)")
def reset_store():
    _store["factures"] = {}
    _store["bons"] = {}
    return {"message": "Store réinitialisé."}


# ---------------------------------------------------------------------------
# Automatisation — pilotage tâches + logs
# ---------------------------------------------------------------------------

@app.get("/api/automation/tasks", summary="Lister les tâches d'automatisation")
def automation_list_tasks():
    with _automation_lock:
        tasks = [dict(v) for v in _automation_tasks.values()]
    tasks.sort(key=lambda x: x.get("id", ""))
    return tasks


@app.get("/api/automation/logs", summary="Lister les logs d'automatisation")
def automation_list_logs(task_id: str | None = None, limit: int = 200):
    lim = max(1, min(limit, 1000))
    with _automation_lock:
        logs = list(_automation_logs)
    if task_id:
        logs = [l for l in logs if l.get("task_id") == task_id]
    return logs[-lim:]


@app.post("/api/automation/tasks/{task_id}/start", summary="Activer une tâche d'automatisation")
def automation_start_task(task_id: str):
    with _automation_lock:
        task = _automation_tasks.get(task_id)
        if not task:
            raise HTTPException(status_code=404, detail=f"Tache '{task_id}' introuvable")
        task["enabled"] = True
        task["next_run"] = datetime.now().isoformat(timespec="seconds")
    _add_automation_log(task_id, "info", "Tache activee.")
    return {"message": f"Tache '{task_id}' activee."}


@app.post("/api/automation/tasks/{task_id}/stop", summary="Désactiver une tâche d'automatisation")
def automation_stop_task(task_id: str):
    with _automation_lock:
        task = _automation_tasks.get(task_id)
        if not task:
            raise HTTPException(status_code=404, detail=f"Tache '{task_id}' introuvable")
        task["enabled"] = False
    _add_automation_log(task_id, "warn", "Tache desactivee.")
    return {"message": f"Tache '{task_id}' desactivee."}


@app.post("/api/automation/tasks/{task_id}/run-now", summary="Exécuter une tâche immédiatement")
def automation_run_task_now(task_id: str):
    with _automation_lock:
        task = _automation_tasks.get(task_id)
        if not task:
            raise HTTPException(status_code=404, detail=f"Tache '{task_id}' introuvable")
        if task.get("is_running"):
            raise HTTPException(status_code=409, detail=f"La tache '{task_id}' est deja en cours")
    _executor.submit(_execute_automation_task, task_id, "manual")
    return {"message": f"Execution immediate lancee pour '{task_id}'."}


# ---------------------------------------------------------------------------
# DOMINO — Automatisation import rapport journalier
# ---------------------------------------------------------------------------

@app.get("/api/domino/files", summary="Lister les fichiers DOMINO et leur statut d'import")
def domino_list_files():
    """Liste les fichiers .xlsx du dossier test_domino/ avec leur statut."""
    return domino_module.list_domino_files()


@app.get("/api/domino/data", summary="Données DOMINO importées")
def domino_get_data():
    """Retourne toutes les données DOMINO importées, triées par date décroissante."""
    return domino_module.get_all_imported_data()


@app.post("/api/domino/import-json", summary="Importer un JSON DOMINO (robuste)")
async def domino_import_json(file: UploadFile = File(...), mode: str = "merge"):
    """
    Importe un JSON DOMINO avec validation/normalisation.
    mode=merge (défaut) ou mode=replace.
    """
    try:
        content = await file.read()
        payload = json.loads(content.decode("utf-8"))
        result = domino_module.import_json_payload(payload, mode=mode)
        return result
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=422, detail=f"JSON invalide: {e}")
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur import JSON DOMINO : {e}")


@app.post("/api/domino/import/{filename}", summary="Importer un fichier DOMINO")
def domino_import_file(filename: str, overwrite: bool = False):
    """
    Parse et importe un fichier DOMINO depuis test_domino/.
    Si overwrite=false (défaut) et que le fichier a déjà été importé, retourne skipped=true.
    Tente d'écrire dans l'onglet DOMINO du XLSM si disponible.
    """
    active_xlsm = _pick_valid_tresorerie_path()

    try:
        with _xlsm_write_lock:
            result = domino_module.import_domino_file(
                filename=filename,
                xlsm_path=active_xlsm,
                overwrite=overwrite,
            )
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur import DOMINO : {e}")

    return result


@app.post("/api/domino/import-all", summary="Importer tous les fichiers DOMINO non encore traités")
def domino_import_all(overwrite: bool = False):
    """
    Importe en batch tous les fichiers DOMINO du dossier test_domino/.
    Ignore par défaut les fichiers déjà importés.
    """
    files = domino_module.list_domino_files()
    if not overwrite:
        files = [f for f in files if not f["imported"]]

    if not files:
        return {"message": "Aucun fichier à importer.", "results": []}

    active_xlsm = _pick_valid_tresorerie_path()
    results = []
    for f in files:
        try:
            with _xlsm_write_lock:
                result = domino_module.import_domino_file(
                    filename=f["filename"],
                    xlsm_path=active_xlsm,
                    overwrite=overwrite,
                )
            results.append(result)
        except Exception as e:
            results.append({
                "filename": f["filename"],
                "date": f.get("date"),
                "skipped": False,
                "xlsm_updated": False,
                "cells_written": 0,
                "message": f"Erreur : {e}",
            })

    imported = sum(1 for r in results if not r.get("skipped") and not r.get("xlsm_error") and "Erreur" not in r.get("message", ""))
    return {
        "message": f"{imported}/{len(results)} fichier(s) importé(s).",
        "results": results,
    }


@app.post("/api/domino/resync-xlsm", summary="Forcer la resynchronisation DOMINO XLSM depuis le JSON")
def domino_resync_xlsm_from_json(force_overwrite: bool = True):
    """
    Rejoue toutes les données DOMINO déjà présentes en JSON vers l'onglet DOMINO du XLSM.
    Par défaut, écrase les valeurs existantes de la colonne date (force_overwrite=true).
    """
    active_xlsm = _ensure_valid_tresorerie_path()

    try:
        with _xlsm_write_lock:
            result = domino_module.resync_xlsm_from_json(
                xlsm_path=active_xlsm,
                force_overwrite=force_overwrite,
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur resynchronisation DOMINO : {e}")

    result["fichier"] = os.path.basename(active_xlsm)
    return result


def _run_domino_resync_job(job_id: str, force_overwrite: bool) -> None:
    """Worker de fond pour la resynchronisation DOMINO JSON -> XLSM."""
    with _domino_resync_jobs_lock:
        _domino_resync_jobs[job_id] = {
            "job_id": job_id,
            "status": "running",
            "message": "Resynchronisation en cours...",
        }

    try:
        active_xlsm = _ensure_valid_tresorerie_path()
        with _xlsm_write_lock:
            result = domino_module.resync_xlsm_from_json(
                xlsm_path=active_xlsm,
                force_overwrite=force_overwrite,
            )
        result["fichier"] = os.path.basename(active_xlsm)
        with _domino_resync_jobs_lock:
            _domino_resync_jobs[job_id] = {
                "job_id": job_id,
                "status": "completed",
                "message": result.get("message", "Resynchronisation terminee."),
                "result": result,
            }
    except Exception as e:
        with _domino_resync_jobs_lock:
            _domino_resync_jobs[job_id] = {
                "job_id": job_id,
                "status": "failed",
                "message": f"Echec de la resynchronisation: {e}",
                "error": str(e),
            }


@app.post("/api/domino/resync-xlsm/start", summary="Démarrer une resynchronisation DOMINO en tâche de fond")
def domino_resync_xlsm_start(force_overwrite: bool = True):
    """
    Lance la resynchronisation en arrière-plan et retourne immédiatement un job_id.
    Évite les timeouts proxy sur les gros fichiers XLSM.
    """
    job_id = str(uuid.uuid4())
    _executor.submit(_run_domino_resync_job, job_id, force_overwrite)
    return {
        "job_id": job_id,
        "status": "running",
        "message": "Resynchronisation DOMINO demarree.",
    }


@app.get("/api/domino/resync-xlsm/status/{job_id}", summary="Statut d'un job de resynchronisation DOMINO")
def domino_resync_xlsm_status(job_id: str):
    with _domino_resync_jobs_lock:
        job = _domino_resync_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job '{job_id}' introuvable")
    return job


# ---------------------------------------------------------------------------
# Helpers internes
# ---------------------------------------------------------------------------

def _deserialize_record(record: dict) -> dict:
    """Reconvertit les strings ISO en objets date pour les traitements internes."""
    out = dict(record)
    for key in ("date_emission", "date_paiement_prevue", "date_livraison"):
        val = out.get(key)
        if isinstance(val, str):
            try:
                out[key] = date.fromisoformat(val)
            except ValueError:
                pass
    return out


def _relink_store() -> None:
    """
    Relance la liaison automatique BL ↔ Factures sur l'ensemble du store
    et réindexe. Préserve les rattachements manuels existants.
    """
    factures_list = list(_store["factures"].values())
    bons_list     = list(_store["bons"].values())
    factures_linked, bons_linked = link_documents(factures_list, bons_list)

    _store["factures"] = {
        f["numero_facture"]: f
        for f in factures_linked
        if f.get("numero_facture")
    }
    _store["bons"] = {
        b["numero_bon_livraison"]: b
        for b in bons_linked
        if b.get("numero_bon_livraison")
    }


def _regenerate_excel():
    """
    Persiste le store dans l'onglet 'Achats Cons' du fichier
    'Suivi trésorerie MLC.xlsm' en écrasant uniquement les lignes
    dont le fournisseur est Sysco / Ambelys / TerreAzur (col C).
    Les lignes des autres fournisseurs sont conservées intactes.
    """
    active_xlsm = _pick_valid_tresorerie_path()
    if not active_xlsm:
        print(
            "[WARN] _regenerate_excel : aucun xlsm valide, persistance ignoree. "
            f"Etat: {_build_tresorerie_invalid_detail()}"
        )
        return

    try:
        factures = list(_store["factures"].values())
        bons     = list(_store["bons"].values())
        with _xlsm_write_lock:
            write_to_achats_cons(
                factures=factures,
                bons=bons,
                template_path=active_xlsm,
                output_path=active_xlsm,
                fournisseur_display=_get_fournisseur_display(),
            )
    except Exception as e:
        err = f"{type(e).__name__}: {e}".rstrip(": ")
        print(f"[WARN] _regenerate_excel : erreur lors de l'ecriture dans le xlsm : {err}")


def _schedule_regenerate_excel() -> None:
    """
    Déclenche une persistance XLSM en arrière-plan.
    Les demandes concurrentes sont fusionnées pour éviter les blocages en rafale.
    """
    global _regen_pending, _regen_running

    with _regen_lock:
        _regen_pending = True
        if _regen_running:
            return
        _regen_running = True

    def _worker():
        global _regen_pending, _regen_running
        while True:
            with _regen_lock:
                if not _regen_pending:
                    _regen_running = False
                    return
                _regen_pending = False
            _regenerate_excel()

    threading.Thread(target=_worker, daemon=True).start()
