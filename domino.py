"""
Module d'extraction et d'import des données DOMINO.

Chaque matin (~4h), un fichier YYYYMMDD.xlsx est déposé dans DOMINO_FOLDER.
Ce module :
  1. Parse le rapport journalier Bassin à Flot
  2. Stocke les imports dans DOMINO_IMPORTS_FILE (JSON)
  3. Écrit les données dans l'onglet DOMINO du fichier Suivi trésorerie MLC.xlsm

Mapping shifts source → lignes DOMINO :
  - CA TTC client matin     (row 7)  → 0 (pas de shift dédié)
  - CA TTC client midi      (row 8)  → shift "MIDI (10-15)"
  - CA TTC client après-midi(row 9)  → shift "04h-10h" (si présent en col C)
  - CA TTC client soir      (row 10) → shift "SOIR (18:30-Fermeture)"
"""

from __future__ import annotations

import json
import os
import re
import tempfile
from dataclasses import asdict, dataclass
from datetime import date, datetime
from typing import Optional, Any

import openpyxl
from xlsm_safe import atomic_save_workbook
import db
import repositories as repo

DOMINO_FOLDER = "test_domino"
DOMINO_IMPORTS_FILE = "output/domino_imports.json"
DOMINO_IMPORTS_BACKUP = "output/domino_imports.lastgood.bak.json"


# ---------------------------------------------------------------------------
# Modèle de données
# ---------------------------------------------------------------------------

@dataclass
class DominoJourData:
    date: date
    filename: str

    # CA TTC par service (boutique)
    ca_ttc_matin: float = 0.0        # row 7  — pas de shift dédié dans la source
    ca_ttc_midi: float = 0.0         # row 8  — shift MIDI (10-15)
    ca_ttc_apm: float = 0.0          # row 9  — shift 04h-10h (si présent)
    ca_ttc_soir: float = 0.0         # row 10 — shift SOIR

    # CA TTC canaux LAD
    ca_ttc_uber: float = 0.0         # row 11
    ca_ttc_deliveroo: float = 0.0    # row 12

    # CA TTC total
    ca_ttc_total: float = 0.0        # row 36

    # TVA collectée
    tva_total: float = 0.0           # row 18
    tva_55: float = 0.0              # row 19
    tva_10: float = 0.0              # row 20

    # Modes de paiement
    especes: float = 0.0             # row 38
    carte_bancaire: float = 0.0      # row 39
    cb_link: float = 0.0             # row 43
    belorder: float = 0.0            # row 45
    uber_eats: float = 0.0           # row 46
    deliveroo_paiement: float = 0.0  # row 47
    total_encaissements: float = 0.0 # row 54

    # Couverts
    nb_clients_matin: int = 0        # row 58
    nb_clients_midi: int = 0         # row 59
    nb_clients_soir: int = 0         # row 61
    total_clients: int = 0           # row 62

    def to_dict(self) -> dict:
        d = asdict(self)
        d["date"] = self.date.isoformat()
        return d


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get(rows: list[tuple], row_idx: int, col_idx: int, default=None):
    if row_idx < 0 or row_idx >= len(rows):
        return default
    row = rows[row_idx]
    if col_idx < 0 or col_idx >= len(row):
        return default
    v = row[col_idx]
    return v if v is not None else default


def _to_float(v, default: float = 0.0) -> float:
    if v is None:
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _to_int(v, default: int = 0) -> int:
    if v is None:
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _parse_date_from_filename(filename: str) -> Optional[date]:
    m = re.match(r"(\d{4})(\d{2})(\d{2})", os.path.basename(filename))
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return None


def _parse_date_from_str(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", v.strip())
        if m:
            try:
                return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            except ValueError:
                pass
    return None


def _find_section_row(rows: list[tuple], keyword: str) -> int:
    """Trouve l'index 0-based de la ligne dont la col A contient keyword."""
    for i, row in enumerate(rows):
        if row and row[0] and isinstance(row[0], str) and keyword.lower() in row[0].lower():
            return i
    return -1


# ---------------------------------------------------------------------------
# Parsing du rapport journalier
# ---------------------------------------------------------------------------

def parse_domino_file(filepath: str) -> DominoJourData:
    """Parse un fichier Rapport Synthèse Journalier Bassin à Flot."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows: list[tuple] = [
        tuple(cell.value for cell in row)
        for row in ws.iter_rows()
    ]
    wb.close()

    # ---- Date ----
    report_date = _parse_date_from_filename(filepath)
    if report_date is None:
        cell_v = _get(rows, 1, 0)  # row 2 col A : "DD/MM/YYYY - DD/MM/YYYY"
        if cell_v:
            report_date = _parse_date_from_str(str(cell_v).split(" - ")[0].strip())
    if report_date is None:
        raise ValueError(f"Impossible de déterminer la date du fichier '{filepath}'")

    data = DominoJourData(date=report_date, filename=os.path.basename(filepath))

    # ---- Section Ventes ----
    v_idx = _find_section_row(rows, "Ventes")
    if v_idx >= 0:
        # Chercher la ligne "HD Ticket | Pourcentage TVA | ..."
        hdr = next(
            (i for i in range(v_idx, min(v_idx + 6, len(rows)))
             if rows[i] and rows[i][0] and "HD Ticket" in str(rows[i][0])),
            -1,
        )
        if hdr >= 0:
            d_row = hdr + 1  # ligne de données totale (date en col A)
            data.ca_ttc_total = _to_float(_get(rows, d_row, 4))
            data.tva_total    = _to_float(_get(rows, d_row, 6))

            # Lignes suivantes : col B = taux TVA (5.5 / 10)
            for i in range(d_row + 1, min(d_row + 12, len(rows))):
                taux = _get(rows, i, 1)
                if taux is None:
                    continue
                taux_f = _to_float(taux)
                tva_v  = _to_float(_get(rows, i, 6))
                if abs(taux_f - 5.5) < 0.1:
                    data.tva_55 = tva_v
                elif abs(taux_f - 10.0) < 0.1:
                    data.tva_10 = tva_v

    # ---- Section Paiements ----
    p_idx = _find_section_row(rows, "Paiements")
    if p_idx >= 0:
        hdr = next(
            (i for i in range(p_idx, min(p_idx + 6, len(rows)))
             if rows[i] and rows[i][0] and "PaymentDT" in str(rows[i][0])),
            -1,
        )
        if hdr >= 0:
            total_row = hdr + 1
            data.total_encaissements = _to_float(_get(rows, total_row, 4))

            for i in range(total_row + 1, min(total_row + 20, len(rows))):
                mode = _get(rows, i, 1)
                if mode is None:
                    break
                mode_s = str(mode).strip().lower()
                val = _to_float(_get(rows, i, 4))
                if "espece" in mode_s:
                    data.especes = val
                elif "carte bancaire" in mode_s or "carte_bancaire" in mode_s:
                    data.carte_bancaire = val
                elif "deliveroo" in mode_s:
                    data.deliveroo_paiement = val
                elif "cb link" in mode_s or "cb_link" in mode_s:
                    data.cb_link = val
                elif "belorder" in mode_s:
                    data.belorder = val
                elif "uber" in mode_s:
                    data.uber_eats = val

        data.ca_ttc_uber      = data.uber_eats
        data.ca_ttc_deliveroo = data.deliveroo_paiement

    # ---- Section Analyse Couverts ----
    c_idx = _find_section_row(rows, "Analyse Couverts")
    if c_idx >= 0:
        hdr = next(
            (i for i in range(c_idx, min(c_idx + 6, len(rows)))
             if rows[i] and rows[i][0] and "HD Ticket" in str(rows[i][0])),
            -1,
        )
        if hdr >= 0:
            total_row = hdr + 1
            data.total_clients = _to_int(_get(rows, total_row, 5))

            for i in range(total_row + 1, min(total_row + 10, len(rows))):
                shift = _get(rows, i, 2)
                if shift is None:
                    continue
                shift_s = str(shift).strip().lower()
                ca_v  = _to_float(_get(rows, i, 4))
                nb_v  = _to_int(_get(rows, i, 5))
                if "04h" in shift_s:
                    # 04h-10h → après-midi dans le DOMINO (si présent en col C)
                    data.ca_ttc_apm    = ca_v
                    data.nb_clients_matin = nb_v  # couverts matin = 04h-10h
                elif "midi" in shift_s or "10-15" in shift_s:
                    data.ca_ttc_midi   = ca_v
                    data.nb_clients_midi = nb_v
                elif "soir" in shift_s or "18" in shift_s:
                    data.ca_ttc_soir   = ca_v
                    data.nb_clients_soir = nb_v

    return data


# ---------------------------------------------------------------------------
# Gestion du dossier et du fichier d'import
# ---------------------------------------------------------------------------

def list_domino_files() -> list[dict]:
    """Liste les fichiers .xlsx dans DOMINO_FOLDER avec leur statut d'import depuis la BDD."""
    # Charger les dates importées depuis la BDD
    imported_dates = {row.get("date"): row for row in repo.list_domino_jours()}
    
    result = []
    if not os.path.exists(DOMINO_FOLDER):
        return result
    for fname in sorted(os.listdir(DOMINO_FOLDER), reverse=True):
        if not fname.lower().endswith(".xlsx"):
            continue
        d = _parse_date_from_filename(fname)
        date_iso = d.isoformat() if d else None
        is_imported = date_iso and date_iso in imported_dates
        imported_at = None
        if is_imported:
            imported_at = imported_dates[date_iso].get("imported_at")
        
        result.append({
            "filename": fname,
            "date": date_iso,
            "imported": is_imported,
            "imported_at": imported_at,
        })
    return result


def get_all_imported_data() -> list[dict]:
    """Retourne toutes les données importées depuis la BDD, triées par date décroissante."""
    domino_rows = repo.list_domino_jours()
    # Récupérer les jours depuis la BDD
    items = []
    for row in domino_rows:
        items.append({
            "imported_at": row.get("imported_at"),
            "filename": row.get("filename"),
            "data": _row_to_dict(row),
        })
    items.sort(key=lambda x: x.get("data", {}).get("date", ""), reverse=True)
    return items


def has_imported_data() -> bool:
    """Indique si la BDD DOMINO contient au moins une entrée."""
    return len(repo.list_domino_jours()) > 0


# ---------------------------------------------------------------------------
# Helpers pour migration JSON → SQLite
# ---------------------------------------------------------------------------

def _row_to_dict(row: dict) -> dict:
    """Convertit une ligne de la BDD domino_jours en dict compatible JSON domino."""
    return {
        "date": row.get("date"),
        "filename": row.get("filename"),
        "ca_ttc_matin": row.get("ca_ttc_matin"),
        "ca_ttc_midi": row.get("ca_ttc_midi"),
        "ca_ttc_apm": row.get("ca_ttc_apm"),
        "ca_ttc_soir": row.get("ca_ttc_soir"),
        "ca_ttc_uber": row.get("ca_ttc_uber"),
        "ca_ttc_deliveroo": row.get("ca_ttc_deliveroo"),
        "ca_ttc_total": row.get("ca_ttc_total"),
        "tva_total": row.get("tva_total"),
        "tva_55": row.get("tva_55"),
        "tva_10": row.get("tva_10"),
        "especes": row.get("especes"),
        "carte_bancaire": row.get("carte_bancaire"),
        "cb_link": row.get("cb_link"),
        "belorder": row.get("belorder"),
        "uber_eats": row.get("uber_eats"),
        "deliveroo_paiement": row.get("deliveroo_paiement"),
        "total_encaissements": row.get("total_encaissements"),
        "nb_clients_matin": row.get("nb_clients_matin"),
        "nb_clients_midi": row.get("nb_clients_midi"),
        "nb_clients_soir": row.get("nb_clients_soir"),
        "total_clients": row.get("total_clients"),
    }



def _load_imports() -> dict:
    if not os.path.exists(DOMINO_IMPORTS_FILE):
        return {}

    try:
        with open(DOMINO_IMPORTS_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
        return _normalize_imports_payload(raw)
    except (json.JSONDecodeError, OSError):
        # Fallback de robustesse: tenter la backup JSON last-good
        if os.path.exists(DOMINO_IMPORTS_BACKUP):
            try:
                with open(DOMINO_IMPORTS_BACKUP, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                return _normalize_imports_payload(raw)
            except (json.JSONDecodeError, OSError):
                return {}
        return {}


def _atomic_write_json(path: str, payload: dict) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(prefix="domino_imports.", suffix=".tmp", dir=os.path.dirname(path) or ".")
    os.close(fd)
    try:
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        if os.path.exists(path):
            # backup du dernier json valide
            with open(path, "r", encoding="utf-8") as src, open(DOMINO_IMPORTS_BACKUP, "w", encoding="utf-8") as dst:
                dst.write(src.read())
        os.replace(tmp_path, path)
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


def _normalize_imports_payload(raw: Any) -> dict:
    """
    Normalise et valide le payload JSON des imports DOMINO.
    Accepte:
    - dict (format interne: key -> {imported_at, filename, data})
    - list de records (auto-conversion)
    """
    if raw is None:
        return {}

    out: dict[str, dict] = {}

    if isinstance(raw, list):
        for item in raw:
            if not isinstance(item, dict):
                continue
            data = item.get("data") if isinstance(item.get("data"), dict) else item
            d = data.get("date")
            if not isinstance(d, str):
                continue
            key = str(item.get("key") or item.get("filename") or d)
            out[key] = {
                "imported_at": str(item.get("imported_at") or datetime.now().isoformat()),
                "filename": str(item.get("filename") or data.get("filename") or f"{key}.xlsx"),
                "data": data,
            }
        return out

    if isinstance(raw, dict):
        for key, item in raw.items():
            if not isinstance(item, dict):
                continue
            data = item.get("data")
            if not isinstance(data, dict):
                continue
            d = data.get("date")
            if not isinstance(d, str):
                continue
            out[str(key)] = {
                "imported_at": str(item.get("imported_at") or datetime.now().isoformat()),
                "filename": str(item.get("filename") or data.get("filename") or f"{key}.xlsx"),
                "data": data,
            }
        return out

    return {}


def _save_import(key: str, data: DominoJourData) -> None:
    """Sauvegarde l'import DOMINO dans la BDD SQLite (le JSON n'est plus utilisé)."""
    repo.upsert_domino_jour(data)


def import_json_payload(raw: Any, mode: str = "merge") -> dict:
    """
    Importe un payload JSON DOMINO de manière robuste.
    Stocke dans la BDD SQLite (le mode n'affecte que la sémantique de présentation).

    mode:
      - merge (défaut): fusionne avec l'existant
      - replace: remplace totalement (sémantiquement; BDD atomique)
    """
    if not isinstance(raw, (list, dict)):
        raise ValueError("Payload invalide: doit être list ou dict")
    
    normalized = _normalize_imports_payload(raw)
    if mode not in {"merge", "replace"}:
        raise ValueError("mode invalide (attendu: merge|replace)")

    imported = 0
    for key, item in normalized.items():
        data_dict = item.get("data")
        if not data_dict:
            continue
        try:
            domino_data = _data_from_import_dict(data_dict)
            repo.upsert_domino_jour(domino_data)
            imported += 1
        except Exception as e:
            print(f"[WARN] Import JSON DOMINO {key}: {e}")

    total = len(repo.list_domino_jours())
    return {
        "message": "Import JSON DOMINO termine.",
        "mode": mode,
        "imported": imported,
        "total": total,
    }


def is_imported(filename: str) -> bool:
    """Vérifie si une date DOMINO a déjà été importée via la BDD."""
    d = _parse_date_from_filename(filename)
    if d is None:
        return False
    return repo.has_domino_jour(d.isoformat())


# ---------------------------------------------------------------------------
# Écriture dans l'onglet DOMINO du XLSM
# ---------------------------------------------------------------------------

def date_to_excel_serial(d: date) -> int:
    """Convertit une date Python en numéro de série Excel (système 1900)."""
    return (d - date(1899, 12, 30)).days


# Mapping ligne DOMINO → attribut DominoJourData
_DOMINO_ROW_MAP: dict[int, str] = {
    7:  "ca_ttc_matin",
    8:  "ca_ttc_midi",
    9:  "ca_ttc_apm",
    10: "ca_ttc_soir",
    11: "ca_ttc_uber",
    12: "ca_ttc_deliveroo",
    18: "tva_total",
    19: "tva_55",
    20: "tva_10",
    36: "ca_ttc_total",
    38: "especes",
    39: "carte_bancaire",
    43: "cb_link",
    45: "belorder",
    46: "uber_eats",
    47: "deliveroo_paiement",
    54: "total_encaissements",
    58: "nb_clients_matin",
    59: "nb_clients_midi",
    61: "nb_clients_soir",
    62: "total_clients",
}


def write_to_domino_sheet(
    data: DominoJourData,
    xlsm_path: str,
    overwrite: bool = False,
) -> dict:
    """
    Écrit les données d'un jour dans l'onglet DOMINO du XLSM.
    Retourne {"col_found": bool, "col_index": int, "cells_written": int, "skipped": bool}.
    """
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    try:
        if "DOMINO" not in wb.sheetnames:
            raise ValueError("Onglet 'DOMINO' introuvable dans le fichier XLSM.")

        ws = wb["DOMINO"]
        result = _write_data_on_open_sheet(ws, data, overwrite=overwrite)
        if not result.get("skipped"):
            atomic_save_workbook(wb, xlsm_path)
        return result
    finally:
        wb.close()


def _find_target_col_for_date(ws, d: date) -> Optional[int]:
    """Retourne la colonne DOMINO associée à la date (ligne 6 = serial excel)."""
    target_serial = date_to_excel_serial(d)
    for col in range(2, ws.max_column + 2):
        cell_val = ws.cell(row=6, column=col).value
        if cell_val is None:
            continue
        try:
            if int(cell_val) == target_serial:
                return col
        except (TypeError, ValueError):
            pass
    return None


def _write_data_on_open_sheet(ws, data: DominoJourData, overwrite: bool) -> dict:
    """Écrit une journée DOMINO sur un worksheet déjà ouvert (sans save)."""
    target_col = _find_target_col_for_date(ws, data.date)
    if target_col is None:
        target_serial = date_to_excel_serial(data.date)
        raise ValueError(
            f"Date {data.date.strftime('%d/%m/%Y')} (serial={target_serial}) "
            f"introuvable dans la ligne 6 de l'onglet DOMINO."
        )

    # Si pas overwrite, vérifier si la colonne a déjà des données (CA midi ≠ 0)
    if not overwrite:
        existing = ws.cell(row=8, column=target_col).value
        if existing is not None and existing != 0:
            return {
                "col_found": True,
                "col_index": target_col,
                "cells_written": 0,
                "skipped": True,
            }

    cells_written = 0
    data_d = data.to_dict()
    for row_num, attr in _DOMINO_ROW_MAP.items():
        value = data_d.get(attr, 0)
        # Écrire None pour les zéros (laisse les formules existantes)
        ws.cell(row=row_num, column=target_col).value = value if value else None
        cells_written += 1

    return {
        "col_found": True,
        "col_index": target_col,
        "cells_written": cells_written,
        "skipped": False,
    }


def _data_from_import_dict(d: dict) -> DominoJourData:
    """Reconstruit un DominoJourData à partir du JSON d'import."""
    if "date" not in d:
        raise ValueError("Entrée JSON DOMINO invalide: date manquante")

    d_date = d.get("date")
    parsed_date: Optional[date] = None
    if isinstance(d_date, str):
        try:
            parsed_date = date.fromisoformat(d_date)
        except ValueError as e:
            raise ValueError(f"Date JSON DOMINO invalide: {d_date}") from e
    if not parsed_date:
        raise ValueError("Entrée JSON DOMINO invalide: date non parseable")

    return DominoJourData(
        date=parsed_date,
        filename=str(d.get("filename") or "json"),
        ca_ttc_matin=_to_float(d.get("ca_ttc_matin")),
        ca_ttc_midi=_to_float(d.get("ca_ttc_midi")),
        ca_ttc_apm=_to_float(d.get("ca_ttc_apm")),
        ca_ttc_soir=_to_float(d.get("ca_ttc_soir")),
        ca_ttc_uber=_to_float(d.get("ca_ttc_uber")),
        ca_ttc_deliveroo=_to_float(d.get("ca_ttc_deliveroo")),
        ca_ttc_total=_to_float(d.get("ca_ttc_total")),
        tva_total=_to_float(d.get("tva_total")),
        tva_55=_to_float(d.get("tva_55")),
        tva_10=_to_float(d.get("tva_10")),
        especes=_to_float(d.get("especes")),
        carte_bancaire=_to_float(d.get("carte_bancaire")),
        cb_link=_to_float(d.get("cb_link")),
        belorder=_to_float(d.get("belorder")),
        uber_eats=_to_float(d.get("uber_eats")),
        deliveroo_paiement=_to_float(d.get("deliveroo_paiement")),
        total_encaissements=_to_float(d.get("total_encaissements")),
        nb_clients_matin=_to_int(d.get("nb_clients_matin")),
        nb_clients_midi=_to_int(d.get("nb_clients_midi")),
        nb_clients_soir=_to_int(d.get("nb_clients_soir")),
        total_clients=_to_int(d.get("total_clients")),
    )


def resync_xlsm_from_json(xlsm_path: str, force_overwrite: bool = True) -> dict:
    """
    Réécrit l'onglet DOMINO dans le XLSM à partir de la BDD SQLite.
    Remplace l'ancienne implémentation qui lisait depuis JSON.

    Par défaut, force l'écrasement des colonnes dates existantes.
    """
    domino_rows = repo.list_domino_jours()
    if not domino_rows:
        return {
            "message": "Aucune donnée DOMINO en BDD à resynchroniser.",
            "total": 0,
            "written": 0,
            "skipped": 0,
            "errors": [],
        }

    # Trier par date
    domino_rows_sorted = sorted(domino_rows, key=lambda x: x.get("date", ""))

    written = 0
    skipped = 0
    errors: list[dict] = []

    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    try:
        if "DOMINO" not in wb.sheetnames:
            raise ValueError("Onglet 'DOMINO' introuvable dans le fichier XLSM.")
        ws = wb["DOMINO"]

        for row in domino_rows_sorted:
            data_dict = _row_to_dict(row)
            try:
                data_obj = _data_from_import_dict(data_dict)
                result = _write_data_on_open_sheet(ws, data_obj, overwrite=force_overwrite)
                if result.get("skipped"):
                    skipped += 1
                else:
                    written += 1
            except Exception as e:
                errors.append(
                    {
                        "date": data_dict.get("date"),
                        "filename": data_dict.get("filename"),
                        "error": str(e),
                    }
                )

        if written > 0:
            atomic_save_workbook(wb, xlsm_path)
    finally:
        wb.close()

    return {
        "message": f"Resynchronisation DOMINO terminee: {written} ecritures, {skipped} ignores, {len(errors)} erreur(s).",
        "total": len(domino_rows),
        "written": written,
        "skipped": skipped,
        "errors": errors,
    }


# ---------------------------------------------------------------------------
# Import complet (parse + save JSON + write XLSM si disponible)
# ---------------------------------------------------------------------------

def import_domino_file(
    filename: str,
    xlsm_path: Optional[str],
    overwrite: bool = False,
) -> dict:
    """
    Importe un fichier DOMINO :
    - Parse le fichier source
    - Sauvegarde dans le JSON de suivi
    - Tente d'écrire dans le XLSM si un chemin valide est fourni

    Retourne un dict résumant l'opération.
    """
    filepath = os.path.join(DOMINO_FOLDER, filename)
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Fichier introuvable : '{filepath}'")

    key = os.path.splitext(filename)[0]
    if not overwrite and is_imported(filename):
        existing = _load_imports().get(key, {})
        return {
            "filename": filename,
            "date": existing.get("data", {}).get("date"),
            "skipped": True,
            "message": "Déjà importé. Utilisez overwrite=true pour forcer.",
            "xlsm_updated": False,
            "cells_written": 0,
        }

    data = parse_domino_file(filepath)
    _save_import(key, data)

    xlsm_result: dict = {"col_found": False, "cells_written": 0, "skipped": False}
    xlsm_updated = False
    xlsm_error: Optional[str] = None

    if xlsm_path:
        try:
            xlsm_result = write_to_domino_sheet(data, xlsm_path, overwrite=overwrite)
            xlsm_updated = xlsm_result.get("cells_written", 0) > 0
        except Exception as e:
            xlsm_error = str(e)

    return {
        "filename": filename,
        "date": data.date.isoformat(),
        "skipped": False,
        "xlsm_updated": xlsm_updated,
        "cells_written": xlsm_result.get("cells_written", 0),
        "xlsm_error": xlsm_error,
        "message": (
            f"Import réussi. {xlsm_result.get('cells_written', 0)} cellules écrites dans DOMINO."
            if xlsm_updated
            else (
                f"Données sauvegardées. XLSM non mis à jour : {xlsm_error}"
                if xlsm_error
                else "Données sauvegardées (XLSM indisponible)."
            )
        ),
        "data": data.to_dict(),
    }
